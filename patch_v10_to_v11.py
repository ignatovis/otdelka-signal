"""
Минимальный патч v10 → v11:
1. СПРАВОЧНИК: вставить столбец Чист:Мебл. (c16), сдвинуть ∑ и остальное
2. ЧИСТ листы: расширить до 5 поверхностей на категорию (Пол/Стены/Потолок/Запол./Мебл.)
3. SIGNAL: обновить формулы под новую раскладку столбцов ЧИСТ
4. ОТЧЕТ: перестроить с 5 поверхностями в чистовой секции
Всё остальное — без изменений.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as CL
from copy import copy
import os, shutil

DIR = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(DIR, '02_Финальная версия - Отделка для SIGNAL v10.xlsx')
DST = os.path.join(DIR, '03_Финальная версия - Отделка для SIGNAL v11.xlsx')

CHIST_SURFS = ['Пол', 'Стены', 'Потолок', 'Запол.', 'Мебл.']
CHERN_SURFS = ['Пол', 'Стены', 'Потолок']
NI = 5
NC = 3

PODZEM_CHIST_V10 = [
    (2,  4, ['Пол','Стены','Потолок','Запол.']),
    (6,  4, ['Пол','Стены','Потолок','Запол.']),
    (10, 3, ['Пол','Стены','Потолок']),
    (13, 4, ['Пол','Стены','Потолок','Запол.']),
    (17, 4, ['Пол','Стены','Потолок','Запол.']),
    (21, 3, ['Пол','Стены','Потолок']),
]
NADZEM_CHIST_V10 = [
    (2,  4, ['Пол','Стены','Потолок','Мебл.']),
    (6,  4, ['Пол','Стены','Потолок','Запол.']),
    (10, 3, ['Пол','Стены','Потолок']),
    (13, 4, ['Пол','Стены','Потолок','Запол.']),
    (17, 4, ['Пол','Стены','Потолок','Запол.']),
    (21, 3, ['Пол','Стены','Потолок']),
    (24, 3, ['Пол','Стены','Потолок']),
]

SURF_INDEX = {'Пол':0, 'Стены':1, 'Потолок':2, 'Запол.':3, 'Мебл.':4}

PODZEM_CATS = [
    ('8.1.1','Отделка паркинг и рампы'),
    ('8.1.2','Отделка эвакуац. лестн. клетки подзем.'),
    ('8.1.3','Лифт. холлы, тамбур-шлюзы подзем.'),
    ('8.1.4','Технич. помещения подзем.'),
    ('8.1.5','Прочие помещения подзем.'),
    ('8.1.6','Коммерч. помещения подзем.'),
]
NADZEM_CATS = [
    ('8.2.1','Отделка лобби/гранд-лобби'),
    ('8.2.2','Отделка эвакуац. лестн. клетки надзем.'),
    ('8.2.3','Лифт. холлы, тамбур-шлюзы надзем.'),
    ('8.2.4','Технич. помещения надзем.'),
    ('8.2.5','Прочие помещения надзем.'),
    ('8.2.6','Коммерч. помещения надзем.'),
    ('8.2.7','Паркинг и рампы надзем.'),
]
NP = len(PODZEM_CATS)

thin = Side(style='thin')
brd = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='4472C4')
bold_font = Font(name='Calibri', size=11, bold=True)
date_fmt = 'DD.MM.YYYY'


def hdr(ws, r, c, val):
    cell = ws.cell(r, c, val)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = brd
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return cell


def dcell(ws, r, c, val=None, fmt=None):
    cell = ws.cell(r, c, val)
    cell.border = brd
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell


def read_chist_data(ws, v10_layout, max_row):
    data = []
    for r in range(5, max_row + 1):
        if ws.cell(r, 1).value is None:
            break
        row = []
        for start_col, width, surf_names in v10_layout:
            vals = [0.0] * NI
            for si, sname in enumerate(surf_names):
                idx = SURF_INDEX.get(sname, si)
                v = ws.cell(r, start_col + si).value
                vals[idx] = v if v is not None else 0
            row.append(vals)
        data.append(row)
    return data


def read_chist_dates(ws, max_row):
    dates = []
    for r in range(5, max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            break
        dates.append(v)
    return dates


def rebuild_chist_sheet(ws, cats, nc, v10_layout, part_label):
    data = read_chist_data(ws, v10_layout, ws.max_row)
    dates = read_chist_dates(ws, ws.max_row)
    title = ws.cell(1, 1).value

    total_data_cols = nc * NI
    total_cols = 1 + total_data_cols + NI
    sum_start = 2 + total_data_cols

    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    for r in range(1, ws.max_row + 1):
        for c in range(1, max(ws.max_column + 1, total_cols + 1)):
            cell = ws.cell(r, c)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = 'General'

    cell = ws.cell(1, 1, title)
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='4472C4')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    hdr(ws, 2, 1, 'Дата')
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    hdr(ws, 2, 2, part_label)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + total_data_cols)

    for ci in range(nc):
        cc = 2 + ci * NI
        code, name = cats[ci]
        lbl = f"{code} {name}"
        hdr(ws, 3, cc, lbl)
        ws.merge_cells(start_row=3, start_column=cc, end_row=3, end_column=cc + NI - 1)

    hdr(ws, 3, sum_start, 'Итого')
    ws.merge_cells(start_row=3, start_column=sum_start, end_row=3, end_column=sum_start + NI - 1)

    for ci in range(nc):
        for si in range(NI):
            hdr(ws, 4, 2 + ci * NI + si, CHIST_SURFS[si])
    for si in range(NI):
        hdr(ws, 4, sum_start + si, CHIST_SURFS[si])

    nw = len(dates)
    for w in range(nw):
        r = 5 + w
        dcell(ws, r, 1, dates[w], date_fmt)

        for ci in range(nc):
            for si in range(NI):
                val = None
                if w < len(data) and ci < len(data[w]):
                    v = data[w][ci][si]
                    val = v if v and v != 0 else None
                dcell(ws, r, 2 + ci * NI + si, val)

        for si in range(NI):
            parts = [f'{CL(2 + ci * NI + si)}{r}' for ci in range(nc)]
            dcell(ws, r, sum_start + si).value = '=' + '+'.join(parts)

    ws.column_dimensions['A'].width = 14
    for c in range(2, total_cols + 1):
        ws.column_dimensions[CL(c)].width = 10


def rebuild_signal(wb, ws):
    NN = len(NADZEM_CATS)
    nw = 48
    for r in range(14, ws.max_row + 1):
        if ws.cell(r, 2).value is None and ws.cell(r, 3).value is None:
            nw = r - 14
            break
    if nw <= 0:
        nw = 48

    def bc(idx):
        return 2 + idx * 6

    CW_COLS = ['H', 'I', 'J']
    HW_COLS = ['L', 'M', 'N', 'O', 'P']

    def weighted_formula(plan_sheet, fact_sheet, cats, ns, w_cols, sprav_start_row, data_row):
        nc_ = len(cats)
        plan_parts = []
        fact_parts = []
        for ci in range(nc_):
            sr = sprav_start_row + ci
            for si in range(ns):
                dc = CL(2 + ci * ns + si)
                wc = w_cols[si]
                plan_parts.append(f"'{plan_sheet}'!{dc}{data_row}*СПРАВОЧНИК!${wc}${sr}")
                fact_parts.append(f"('{fact_sheet}'!{dc}{data_row})*СПРАВОЧНИК!${wc}${sr}")
        return '=' + '+'.join(plan_parts), '=' + '+'.join(fact_parts)

    def combined_formula(plan_chern, fact_chern, plan_chist, fact_chist,
                         cats, sprav_start, data_row):
        plan_parts = []
        fact_parts = []
        for ci in range(len(cats)):
            sr = sprav_start + ci
            wc = f"СПРАВОЧНИК!$E${sr}"
            wh = f"СПРАВОЧНИК!$F${sr}"
            cp, cf = [], []
            for si in range(NC):
                dc = CL(2 + ci * NC + si)
                sw = CW_COLS[si]
                cp.append(f"'{plan_chern}'!{dc}{data_row}*СПРАВОЧНИК!${sw}${sr}")
                cf.append(f"'{fact_chern}'!{dc}{data_row}*СПРАВОЧНИК!${sw}${sr}")
            ip, iff = [], []
            for si in range(NI):
                dc = CL(2 + ci * NI + si)
                sw = HW_COLS[si]
                ip.append(f"'{plan_chist}'!{dc}{data_row}*СПРАВОЧНИК!${sw}${sr}")
                iff.append(f"'{fact_chist}'!{dc}{data_row}*СПРАВОЧНИК!${sw}${sr}")
            plan_parts.append(f"{wc}*({'+'.join(cp)})+{wh}*({'+'.join(ip)})")
            fact_parts.append(f"{wc}*({'+'.join(cf)})+{wh}*({'+'.join(iff)})")
        return '=' + '+'.join(plan_parts), '=' + '+'.join(fact_parts)

    korpus_names = ['К1', 'К2', 'К3']

    # Clear stale v10 data in all 18 blocks × 6 cols
    for r in range(14, 14 + nw):
        for c in range(2, 2 + 18 * 6):
            ws.cell(r, c).value = None

    for w in range(nw):
        r = 14 + w
        dr = 5 + w

        b0, b1, b2 = bc(0), bc(1), bc(2)

        for b_ in [b0, b1, b2]:
            if w == 0:
                ws.cell(r, b_).value = '=$A$2'
            else:
                ws.cell(r, b_).value = f'={CL(b_)}{r-1}+7'
            ws.cell(r, b_).number_format = date_fmt

        pf, ff = weighted_formula('ПЛАН ЧЕРН ПОДЗЕМ', 'ФАКТ ЧЕРН ПОДЗЕМ',
                                  PODZEM_CATS, NC, CW_COLS, 3, dr)
        ws.cell(r, b1+1).value = pf
        ws.cell(r, b1+2).value = ff

        pf, ff = weighted_formula('ПЛАН ЧИСТ ПОДЗЕМ', 'ФАКТ ЧИСТ ПОДЗЕМ',
                                  PODZEM_CATS, NI, HW_COLS, 3, dr)
        ws.cell(r, b2+1).value = pf
        ws.cell(r, b2+2).value = ff

        pf, ff = combined_formula('ПЛАН ЧЕРН ПОДЗЕМ', 'ФАКТ ЧЕРН ПОДЗЕМ',
                                  'ПЛАН ЧИСТ ПОДЗЕМ', 'ФАКТ ЧИСТ ПОДЗЕМ',
                                  PODZEM_CATS, 3, dr)
        ws.cell(r, b0+1).value = pf
        ws.cell(r, b0+2).value = ff

        kor_o_bcs, kor_c_bcs, kor_i_bcs = [], [], []
        for ki, k in enumerate(korpus_names):
            b_o, b_c, b_i = bc(3+ki*3), bc(4+ki*3), bc(5+ki*3)
            kor_o_bcs.append(b_o)
            kor_c_bcs.append(b_c)
            kor_i_bcs.append(b_i)

            for b_ in [b_o, b_c, b_i]:
                if w == 0:
                    ws.cell(r, b_).value = '=$A$2'
                else:
                    ws.cell(r, b_).value = f'={CL(b_)}{r-1}+7'
                ws.cell(r, b_).number_format = date_fmt

            pf, ff = weighted_formula(f'ПЛАН ЧЕРН {k}', f'ФАКТ ЧЕРН {k}',
                                      NADZEM_CATS, NC, CW_COLS, 3+NP, dr)
            ws.cell(r, b_c+1).value = pf
            ws.cell(r, b_c+2).value = ff

            pf, ff = weighted_formula(f'ПЛАН ЧИСТ {k}', f'ФАКТ ЧИСТ {k}',
                                      NADZEM_CATS, NI, HW_COLS, 3+NP, dr)
            ws.cell(r, b_i+1).value = pf
            ws.cell(r, b_i+2).value = ff

            pf, ff = combined_formula(f'ПЛАН ЧЕРН {k}', f'ФАКТ ЧЕРН {k}',
                                      f'ПЛАН ЧИСТ {k}', f'ФАКТ ЧИСТ {k}',
                                      NADZEM_CATS, 3+NP, dr)
            ws.cell(r, b_o+1).value = pf
            ws.cell(r, b_o+2).value = ff

        b_ano, b_anc, b_ani = bc(12), bc(13), bc(14)
        for b_ in [b_ano, b_anc, b_ani]:
            if w == 0:
                ws.cell(r, b_).value = '=$A$2'
            else:
                ws.cell(r, b_).value = f'={CL(b_)}{r-1}+7'
            ws.cell(r, b_).number_format = date_fmt

        # Weighted by korpus: СПРАВОЧНИК!$K$28(К1), $K$29(К2), $K$30(К3)
        kor_w = [f"СПРАВОЧНИК!$K${28+ki}" for ki in range(len(korpus_names))]
        ws.cell(r, b_anc+1).value = '=' + '+'.join(f'{CL(b+1)}{r}*{kor_w[ki]}' for ki, b in enumerate(kor_c_bcs))
        ws.cell(r, b_anc+2).value = '=' + '+'.join(f'{CL(b+2)}{r}*{kor_w[ki]}' for ki, b in enumerate(kor_c_bcs))
        ws.cell(r, b_ani+1).value = '=' + '+'.join(f'{CL(b+1)}{r}*{kor_w[ki]}' for ki, b in enumerate(kor_i_bcs))
        ws.cell(r, b_ani+2).value = '=' + '+'.join(f'{CL(b+2)}{r}*{kor_w[ki]}' for ki, b in enumerate(kor_i_bcs))
        ws.cell(r, b_ano+1).value = '=' + '+'.join(f'{CL(b+1)}{r}*{kor_w[ki]}' for ki, b in enumerate(kor_o_bcs))
        ws.cell(r, b_ano+2).value = '=' + '+'.join(f'{CL(b+2)}{r}*{kor_w[ki]}' for ki, b in enumerate(kor_o_bcs))

        b_to, b_tc, b_ti = bc(15), bc(16), bc(17)
        for b_ in [b_to, b_tc, b_ti]:
            if w == 0:
                ws.cell(r, b_).value = '=$A$2'
            else:
                ws.cell(r, b_).value = f'={CL(b_)}{r-1}+7'
            ws.cell(r, b_).number_format = date_fmt

        w_pod = 'СПРАВОЧНИК!$W$18'
        w_nad = 'СПРАВОЧНИК!$W$19'
        ws.cell(r, b_tc+1).value = f'={CL(b1+1)}{r}*{w_pod}+{CL(b_anc+1)}{r}*{w_nad}'
        ws.cell(r, b_tc+2).value = f'={CL(b1+2)}{r}*{w_pod}+{CL(b_anc+2)}{r}*{w_nad}'
        ws.cell(r, b_ti+1).value = f'={CL(b2+1)}{r}*{w_pod}+{CL(b_ani+1)}{r}*{w_nad}'
        ws.cell(r, b_ti+2).value = f'={CL(b2+2)}{r}*{w_pod}+{CL(b_ani+2)}{r}*{w_nad}'
        ws.cell(r, b_to+1).value = f'={CL(b0+1)}{r}*{w_pod}+{CL(b_ano+1)}{r}*{w_nad}'
        ws.cell(r, b_to+2).value = f'={CL(b0+2)}{r}*{w_pod}+{CL(b_ano+2)}{r}*{w_nad}'


def rebuild_otchet(wb, ws):
    KORPUS = ['К1', 'К2', 'К3']
    CHERN_W = ['H', 'I', 'J']
    CHIST_W = ['L', 'M', 'N', 'O', 'P']
    CHIST_NAMES = ['Пол', 'Стены', 'Потолок', 'Заполн. проемов', 'Меблировка']

    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    cell = ws.cell(1, 1, 'ОТЧЁТ ПО ОТДЕЛОЧНЫМ РАБОТАМ')
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='4472C4')
    ws.merge_cells('A1:I1')
    ws.cell(2, 1, '⚠ Автоматический. Не вносите данные напрямую.')
    ws.merge_cells('A2:I2')

    row = [4]

    def write_section(section_title, cats, sprav_start, fact_chern_sheet, fact_chist_sheet):
        r = row[0]
        ws.cell(r, 1, section_title)
        ws.cell(r, 1).font = bold_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        r += 1
        for ci_, hd_ in enumerate(['Код','Название','Тип','Вес ур.4','Вес ур.5','План м²','Факт м²','Факт %','Взвеш. %']):
            hdr(ws, r, ci_+1, hd_)
        r += 1

        for ci, (code, name) in enumerate(cats):
            sr = sprav_start + ci
            chern_col_start = 2 + ci * NC
            chist_col_start = 2 + ci * NI

            r_main = r
            r_chern = r + 1
            r_chist = r + 1 + NC + 1

            dcell(ws, r, 1, code)
            dcell(ws, r, 2).value = f"=СПРАВОЧНИК!C{sr}"
            dcell(ws, r, 3, 'Главная')
            dcell(ws, r, 6).value = f"=СПРАВОЧНИК!D{sr}"
            dcell(ws, r, 7).value = f"=G{r_chern}*D{r_chern}+G{r_chist}*D{r_chist}"
            dcell(ws, r, 8).value = f"=IF(F{r}=0,0,G{r}/F{r})"
            dcell(ws, r, 9).value = f"=I{r_chern}+I{r_chist}"
            r += 1

            dcell(ws, r, 2, '  Черновая')
            dcell(ws, r, 3, 'Категория')
            dcell(ws, r, 4).value = f"=СПРАВОЧНИК!E{sr}"
            dcell(ws, r, 6).value = f"=F{r_main}"
            surf_refs = '+'.join(f'G{r+1+s}*E{r+1+s}' for s in range(NC))
            dcell(ws, r, 7).value = f"={surf_refs}"
            dcell(ws, r, 8).value = f"=IF(F{r}=0,0,G{r}/F{r})"
            dcell(ws, r, 9).value = f"=H{r}*D{r}"
            r += 1

            for si in range(NC):
                dcell(ws, r, 2, f'   {CHERN_SURFS[si]}')
                dcell(ws, r, 3, 'Вид')
                dcell(ws, r, 5).value = f"=СПРАВОЧНИК!${CHERN_W[si]}${sr}"
                dcell(ws, r, 6).value = f"=F{r_main}"
                fc = CL(chern_col_start + si)
                dcell(ws, r, 7).value = f"=SUM('{fact_chern_sheet}'!{fc}5:{fc}100)"
                dcell(ws, r, 8).value = f"=IF(F{r}=0,0,G{r}/F{r})"
                dcell(ws, r, 9).value = f"=H{r}*E{r}"
                r += 1

            dcell(ws, r, 2, '  Чистовая')
            dcell(ws, r, 3, 'Категория')
            dcell(ws, r, 4).value = f"=СПРАВОЧНИК!F{sr}"
            dcell(ws, r, 6).value = f"=F{r_main}"
            surf_refs = '+'.join(f'G{r+1+s}*E{r+1+s}' for s in range(NI))
            dcell(ws, r, 7).value = f"={surf_refs}"
            dcell(ws, r, 8).value = f"=IF(F{r}=0,0,G{r}/F{r})"
            dcell(ws, r, 9).value = f"=H{r}*D{r}"
            r += 1

            for si in range(NI):
                dcell(ws, r, 2, f'   {CHIST_NAMES[si]}')
                dcell(ws, r, 3, 'Вид')
                dcell(ws, r, 5).value = f"=СПРАВОЧНИК!${CHIST_W[si]}${sr}"
                dcell(ws, r, 6).value = f"=F{r_main}"
                fc = CL(chist_col_start + si)
                dcell(ws, r, 7).value = f"=SUM('{fact_chist_sheet}'!{fc}5:{fc}100)"
                dcell(ws, r, 8).value = f"=IF(F{r}=0,0,G{r}/F{r})"
                dcell(ws, r, 9).value = f"=H{r}*E{r}"
                r += 1

            dcell(ws, r, 2, '  ✓ Сумма')
            dcell(ws, r, 4).value = f"=D{r_chern}+D{r_chist}"
            r += 2

        row[0] = r

    write_section('ПОДЗЕМНАЯ ЧАСТЬ', PODZEM_CATS, 3,
                  'ФАКТ ЧЕРН ПОДЗЕМ', 'ФАКТ ЧИСТ ПОДЗЕМ')
    for k in KORPUS:
        write_section(f'НАДЗЕМНАЯ ЧАСТЬ — {k}', NADZEM_CATS, 3 + NP,
                      f'ФАКТ ЧЕРН {k}', f'ФАКТ ЧИСТ {k}')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 12
    for c_ in ['D','E','F','G','H','I']:
        ws.column_dimensions[c_].width = 12


def main():
    print('Копирование v10 → v11...')
    shutil.copy2(SRC, DST)

    print('Открытие v11...')
    wb = openpyxl.load_workbook(DST)

    # 1. СПРАВОЧНИК: insert Чист:Мебл. column
    print('СПРАВОЧНИК: вставка столбца Чист:Мебл. (c16)...')
    ws_sprav = wb['СПРАВОЧНИК']
    ws_sprav.insert_cols(16, 1)
    hdr(ws_sprav, 2, 16, 'Чист:Мебл.')

    for r in range(3, 16):
        code = ws_sprav.cell(r, 1).value
        if str(code) == '8.2.1':
            old_val = ws_sprav.cell(r, 15).value or 0
            dcell(ws_sprav, r, 16, old_val)
            dcell(ws_sprav, r, 15, 0)
        else:
            dcell(ws_sprav, r, 16, 0)

    for r in range(3, 16):
        ws_sprav.cell(r, 17).value = f'=SUM(L{r}:P{r})'

    # insert_cols already shifted merges — no manual fix needed

    print('СПРАВОЧНИК: готово (24 столбца)')

    # 2. Rebuild ЧИСТ sheets
    chist_sheets = {
        'ПЛАН ЧИСТ ПОДЗЕМ': (PODZEM_CATS, NP, PODZEM_CHIST_V10, 'Подземная часть'),
        'ФАКТ ЧИСТ ПОДЗЕМ': (PODZEM_CATS, NP, PODZEM_CHIST_V10, 'Подземная часть'),
    }
    for k in ['К1', 'К2', 'К3']:
        chist_sheets[f'ПЛАН ЧИСТ {k}'] = (NADZEM_CATS, len(NADZEM_CATS), NADZEM_CHIST_V10, 'Надземная часть')
        chist_sheets[f'ФАКТ ЧИСТ {k}'] = (NADZEM_CATS, len(NADZEM_CATS), NADZEM_CHIST_V10, 'Надземная часть')

    for sn, (cats, nc, v10_layout, part_label) in chist_sheets.items():
        print(f'{sn}: перестройка (5 поверхностей)...')
        rebuild_chist_sheet(wb[sn], cats, nc, v10_layout, part_label)

    # 3. Rebuild SIGNAL formulas
    print('SIGNAL: обновление формул...')
    rebuild_signal(wb, wb['SIGNAL'])

    # 4. Rebuild ОТЧЕТ
    print('ОТЧЕТ: перестройка с 5 поверхностями...')
    rebuild_otchet(wb, wb['ОТЧЕТ'])

    # 5. ИНСТРУКЦИЯ: fix date description
    ws_instr = wb['ИНСТРУКЦИЯ']
    ws_instr.cell(56, 2).value = (
        'Первая дата — это дата первого отчёта в SIGNAL, '
        'которая входит в период плановых сроков реализации.'
    )

    print(f'Сохранение: {DST}')
    wb.save(DST)
    print('Готово!')


if __name__ == '__main__':
    main()
