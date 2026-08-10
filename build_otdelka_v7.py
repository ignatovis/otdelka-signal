import subprocess
subprocess.run(["pip","install","openpyxl"], capture_output=True)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as cl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime, timedelta

BUILDINGS = ["К1", "К2", "К3"]
N_BLDG = len(BUILDINGS)
wb = openpyxl.Workbook()

BOLD = Font(bold=True)
BW = Font(bold=True, color="FFFFFF", size=11)
B14W = Font(bold=True, size=14, color="FFFFFF")
B12W = Font(bold=True, size=12, color="FFFFFF")
B11 = Font(bold=True, size=11)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LW = Alignment(horizontal="left", vertical="center", wrap_text=True)
BRD = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))

F_TITLE = PatternFill("solid", fgColor="2F5496")
F_UG = PatternFill("solid", fgColor="D9E2F3")
F_AG = PatternFill("solid", fgColor="E2EFDA")
F_TOT = PatternFill("solid", fgColor="FFF2CC")
F_HDR = PatternFill("solid", fgColor="B4C6E7")
F_WARN = PatternFill("solid", fgColor="FFC7CE")
F_OK = PatternFill("solid", fgColor="C6EFCE")
F_CFG = PatternFill("solid", fgColor="D6DCE4")
F_PAR = PatternFill("solid", fgColor="DAEEF3")
F_FORM = PatternFill("solid", fgColor="F2F2F2")
F_ENTRY = PatternFill("solid", fgColor="FFFFFF")
F_PREFILL = PatternFill("solid", fgColor="E8E8E8")
F_PLAN = PatternFill("solid", fgColor="EAF0FB")
F_LVL = PatternFill("solid", fgColor="D9D2E9")
F_LVL1 = PatternFill("solid", fgColor="B4A7D6")
F_BLDG = PatternFill("solid", fgColor="C5E0B4")

def sc(cell, font=None, fill=None, align=None, border=None, nf=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if nf: cell.number_format = nf

def sr(ws, r1, r2, c1, c2, **kw):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            sc(ws.cell(r,c), **kw)

UNDERGROUND = [
    {"code":"8.1.1","name":"Отделка паркинг и рампы","plan":6600,
     "rw":0.5,"fw":0.5,"rs":[0.65,0.34,0.01],"fs":[0.7,0.2,0.1,0]},
    {"code":"8.1.2","name":"Отделка эвакуац. лестн. клетки подзем.","plan":2530,
     "rw":0.4,"fw":0.6,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.1.3","name":"Лифт. холлы, тамбур-шлюзы подзем.","plan":1800,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.1.4","name":"Технич. помещения подзем.","plan":3200,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.1.5","name":"Прочие помещения подзем.","plan":2000,
     "rw":0.5,"fw":0.5,"rs":[0.2,0.6,0.2],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.1.6","name":"Коммерч. помещения подзем.","plan":2500,
     "rw":0.5,"fw":0.5,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
]
ABOVEGROUND = [
    {"code":"8.2.1","name":"Отделка лобби/гранд-лобби","plan":1500,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],
     "fs":[0.15,0.45,0.15,0.25],"has_furn":True},
    {"code":"8.2.2","name":"Отделка эвакуац. лестн. клетки надзем.","plan":2000,
     "rw":0.4,"fw":0.6,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.3","name":"Лифт. холлы, тамбур-шлюзы надзем.","plan":2500,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.4","name":"Технич. помещения надзем.","plan":4430,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.5","name":"Прочие помещения надзем.","plan":4000,
     "rw":0.5,"fw":0.5,"rs":[0.2,0.6,0.2],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.6","name":"Коммерч. помещения надзем.","plan":5000,
     "rw":0.5,"fw":0.5,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.7","name":"Паркинг и рампы надзем.","plan":1500,
     "rw":0.5,"fw":0.5,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
]
ALL = UNDERGROUND + ABOVEGROUND
N_UG = len(UNDERGROUND); N_AG = len(ABOVEGROUND)
SURFS = ["Пол","Стены","Потолок"]

dates = []
d = datetime(2026,6,1)
while d <= datetime(2027,4,30):
    dates.append(d); d += timedelta(days=7)
N_WEEKS = len(dates)

DR = 5; LDR = DR + N_WEEKS - 1
RS = 3; RE = RS + len(ALL) - 1
REF = "СПРАВОЧНИК"; V_DR = 4
VCOLS = {"Пол":"E","Стены":"F","Потолок":"G","Мебл.":"H"}
UG_END = RS + N_UG - 1; AG_START = RS + N_UG
SUM_HDR = RE + 2; SUM_UG = RE + 3; SUM_AG = RE + 4

def rref(cat): return RS + ALL.index(cat)

CFG_COMMENTS = {
    "Статус":"ИСТИНА = карточка активна.\nЛОЖЬ = скрыта.",
    "Гистограмма":"ИСТИНА = гистограмма.\nЛОЖЬ = линейный график.",
    "По месяцам":"ИСТИНА = по месяцам.\nЛОЖЬ = понедельно.",
    "Учитывать дату карточки":"ИСТИНА = фильтр по дате.\nЛОЖЬ = все данные.",
}

def vsum(surf, code, tname, vsheet, vlast, dcell=None):
    sn = f"'{vsheet}'"
    sc_ = VCOLS[surf]
    p = f"{sn}!{sc_}${V_DR}:{sc_}${vlast}"
    cr = ""
    if dcell: cr += f",{sn}!$A${V_DR}:$A${vlast},{dcell}"
    cr += f",{sn}!$C${V_DR}:$C${vlast},\"{code}\""
    cr += f",{sn}!$B${V_DR}:$B${vlast},\"{tname}\""
    return f"SUMIFS({p}{cr})"


# ═══════════════════════════════════════════════════════════
# 1. СПРАВОЧНИК
# ═══════════════════════════════════════════════════════════
ws = wb.active; ws.title = "СПРАВОЧНИК"
ws.sheet_properties.tabColor = "2F5496"

ws.merge_cells("A1:W1")
c = ws["A1"]; c.value = "СПРАВОЧНИК КАТЕГОРИЙ ОТДЕЛКИ"
sc(c,B14W,F_TITLE,CTR); sr(ws,1,1,1,23,fill=F_TITLE)

hdrs = ["Код","Часть","Название","План м²","Вес черн.","Вес чист.","∑ весов",
        "Черн:Пол","Черн:Стены","Черн:Потолок","∑ черн.",
        "Чист:Пол","Чист:Стены","Чист:Потолок","Чист:Мебл.","∑ чист.",
        "Начало","Окончание","Длит.(дни)","Труд-proxy","Вес ур.3 (авто)","Вес ур.3",
        "Глоб. вес"]
for i,h in enumerate(hdrs,1):
    c = ws.cell(2,i,h); sc(c,BOLD,F_HDR,CTR,BRD)

for idx,cat in enumerate(ALL):
    r = RS+idx
    is_ug = cat["code"].startswith("8.1")
    part = "Подзем." if is_ug else "Надзем."
    fl = F_UG if is_ug else F_AG
    vals = [cat["code"],part,cat["name"],cat["plan"],
            cat["rw"],cat["fw"],f"=E{r}+F{r}",
            cat["rs"][0],cat["rs"][1],cat["rs"][2],f"=SUM(H{r}:J{r})",
            cat["fs"][0],cat["fs"][1],cat["fs"][2],cat["fs"][3],f"=SUM(L{r}:O{r})"]
    for ci,v in enumerate(vals,1):
        c = ws.cell(r,ci,v); sc(c,border=BRD,fill=fl)
    for cc in [7,11,16]: ws.cell(r,cc).number_format='0.00'
    sc(ws.cell(r,17),border=BRD,fill=F_ENTRY,nf="DD.MM.YYYY")
    sc(ws.cell(r,18),border=BRD,fill=F_ENTRY,nf="DD.MM.YYYY")
    ws.cell(r,19,f'=IF(OR(Q{r}="",R{r}=""),0,R{r}-Q{r})')
    sc(ws.cell(r,19),border=BRD,fill=F_FORM,nf='#,##0')
    ws.cell(r,20,f"=S{r}*D{r}"); sc(ws.cell(r,20),border=BRD,fill=F_FORM,nf='#,##0')
    ws.cell(r,21,f'=IFERROR(T{r}/SUMIF($B${RS}:$B${RE},$B{r},$T${RS}:$T${RE}),0)')
    sc(ws.cell(r,21),border=BRD,fill=F_FORM,nf='0.0000')
    ws.cell(r,22,f"=U{r}"); sc(ws.cell(r,22),border=BRD,fill=F_ENTRY,nf='0.0000')
    ws.cell(r,23,f'=V{r}*IF($B{r}="Подзем.",$V${SUM_UG},$V${SUM_AG})')
    sc(ws.cell(r,23),border=BRD,fill=F_FORM,nf='0.00%')

ws.cell(2,17).comment=Comment("Введите дату начала работ","Шаблон")
ws.cell(2,18).comment=Comment("Введите дату окончания работ","Шаблон")
ws.cell(2,22).comment=Comment("По умолчанию = авто.\nМожно перезаписать вручную.","Шаблон")

for cr_ in [f"G{RS}:G{RE}",f"K{RS}:K{RE}",f"P{RS}:P{RE}"]:
    ws.conditional_formatting.add(cr_,CellIsRule(operator="notEqual",formula=["1"],fill=F_WARN))
    ws.conditional_formatting.add(cr_,CellIsRule(operator="equal",formula=["1"],fill=F_OK))

dv = DataValidation(type="decimal",operator="between",formula1=0,formula2=1)
ws.add_data_validation(dv); dv.add(f"E{RS}:F{RE}"); dv.add(f"H{RS}:J{RE}"); dv.add(f"L{RS}:O{RE}")
dv2 = DataValidation(type="decimal",operator="greaterThanOrEqual",formula1=0)
ws.add_data_validation(dv2); dv2.add(f"D{RS}:D{RE}")
dv_w = DataValidation(type="decimal",operator="between",formula1=0,formula2=1)
ws.add_data_validation(dv_w); dv_w.add(f"V{RS}:V{RE}")

ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=12; ws.column_dimensions["C"].width=48
ws.column_dimensions["D"].width=14
for x in "EFGHIJKLMNOP": ws.column_dimensions[x].width=13
ws.column_dimensions["Q"].width=14; ws.column_dimensions["R"].width=14
ws.column_dimensions["S"].width=12; ws.column_dimensions["T"].width=14
ws.column_dimensions["U"].width=16; ws.column_dimensions["V"].width=13; ws.column_dimensions["W"].width=12
ws.freeze_panes="A3"
ws.cell(2,15).comment=Comment("Только для 8.2.1 лобби","Шаблон")

# ── Level 2 summary ──
ws.merge_cells(f"A{SUM_HDR}:W{SUM_HDR}")
c = ws.cell(SUM_HDR, 1, "АГРЕГАЦИЯ ПО УРОВНЯМ (ур.3 → ур.2 → ур.1)")
sc(c, B12W, F_TITLE, CTR); sr(ws, SUM_HDR, SUM_HDR, 1, 23, fill=F_TITLE)

for ci,h in {3:"Название",4:"План м²",17:"Начало (мин)",18:"Окончание (макс)",
             19:"Длит.(дни)",20:"Труд-proxy",21:"Вес ур.2 (авто)",22:"Вес ур.2",23:"Глоб. вес"}.items():
    c = ws.cell(SUM_HDR+1, ci, h); sc(c, BOLD, F_HDR, CTR, BRD)

for sr_row, pc, pl, r_s, r_e in [(SUM_UG,"8.1","Подземная часть",RS,UG_END),
                                   (SUM_AG,"8.2","Надземная часть",AG_START,RE)]:
    ws.cell(sr_row,1,pc); sc(ws.cell(sr_row,1),B11,border=BRD)
    ws.cell(sr_row,3,pl); sc(ws.cell(sr_row,3),B11,border=BRD,align=LW)
    ws.cell(sr_row,4,f"=SUM(D{r_s}:D{r_e})"); sc(ws.cell(sr_row,4),BOLD,border=BRD,nf='#,##0')
    ws.cell(sr_row,17,f"=MIN(Q{r_s}:Q{r_e})"); sc(ws.cell(sr_row,17),border=BRD,nf="DD.MM.YYYY")
    ws.cell(sr_row,18,f"=MAX(R{r_s}:R{r_e})"); sc(ws.cell(sr_row,18),border=BRD,nf="DD.MM.YYYY")
    ws.cell(sr_row,19,f"=IF(OR(Q{sr_row}=0,R{sr_row}=0),0,R{sr_row}-Q{sr_row})")
    sc(ws.cell(sr_row,19),border=BRD,nf='#,##0')
    ws.cell(sr_row,20,f"=SUM(T{r_s}:T{r_e})"); sc(ws.cell(sr_row,20),BOLD,border=BRD,nf='#,##0')
    ws.cell(sr_row,21,f"=IFERROR(T{sr_row}/(T{SUM_UG}+T{SUM_AG}),0)")
    sc(ws.cell(sr_row,21),BOLD,border=BRD,nf='0.0000')
    ws.cell(sr_row,22,f"=U{sr_row}"); sc(ws.cell(sr_row,22),BOLD,border=BRD,fill=F_ENTRY,nf='0.0000')
    ws.cell(sr_row,23,f"=V{sr_row}"); sc(ws.cell(sr_row,23),BOLD,border=BRD,nf='0.00%')
    sr(ws, sr_row, sr_row, 1, 23, border=BRD)

SUM_CHK = SUM_AG + 2
for cr,label,formula in [
    (SUM_CHK,  "✓ ∑ вес ур.3 подзем.",f"=SUM(V{RS}:V{UG_END})"),
    (SUM_CHK+1,"✓ ∑ вес ур.3 надзем.",f"=SUM(V{AG_START}:V{RE})"),
    (SUM_CHK+2,"✓ ∑ вес ур.2",       f"=V{SUM_UG}+V{SUM_AG}"),
    (SUM_CHK+3,"✓ ∑ глоб. вес",      f"=SUM(W{RS}:W{RE})"),
]:
    ws.cell(cr,3,label); sc(ws.cell(cr,3),Font(bold=True,italic=True,size=9),align=LW)
    col = 23 if "глоб" in label else 22
    ws.cell(cr,col,formula); sc(ws.cell(cr,col),nf='0.00%' if col==23 else '0.0000')
    ltr = "W" if col==23 else "V"
    ws.conditional_formatting.add(f"{ltr}{cr}",CellIsRule(operator="notEqual",formula=["1"],fill=F_WARN))
    ws.conditional_formatting.add(f"{ltr}{cr}",CellIsRule(operator="equal",formula=["1"],fill=F_OK))

# ── Building matrix ──
BLDG_SEC = SUM_CHK + 5
bm_cols = 2 + N_AG + 2
ws.merge_cells(f"A{BLDG_SEC}:{cl(bm_cols)}{BLDG_SEC}")
c = ws.cell(BLDG_SEC, 1, "РАСПРЕДЕЛЕНИЕ НАДЗЕМНОЙ ЧАСТИ ПО КОРПУСАМ")
sc(c, B12W, F_TITLE, CTR); sr(ws, BLDG_SEC, BLDG_SEC, 1, bm_cols, fill=F_TITLE)

bh = BLDG_SEC + 1
ws.cell(bh,1,"Корпус"); sc(ws.cell(bh,1),BOLD,F_HDR,CTR,BRD)
for j,cat in enumerate(ABOVEGROUND):
    ws.cell(bh,2+j,cat["code"]); sc(ws.cell(bh,2+j),BOLD,F_HDR,CTR,BRD)
SUM_C = 2+N_AG; AW_C = SUM_C+1; W_C = AW_C+1
ws.cell(bh,SUM_C,"Σ план"); sc(ws.cell(bh,SUM_C),BOLD,F_HDR,CTR,BRD)
ws.cell(bh,AW_C,"Вес корп.(авто)"); sc(ws.cell(bh,AW_C),BOLD,F_HDR,CTR,BRD)
ws.cell(bh,W_C,"Вес корп."); sc(ws.cell(bh,W_C),BOLD,F_HDR,CTR,BRD)

BLDG_START = BLDG_SEC + 2
BLDG_ROWS = {}
for bi,bn in enumerate(BUILDINGS):
    br = BLDG_START + bi; BLDG_ROWS[bn] = br
    ws.cell(br,1,bn); sc(ws.cell(br,1),B11,F_BLDG,CTR,BRD)
    for j in range(N_AG):
        ws.cell(br,2+j,ABOVEGROUND[j]["plan"]); sc(ws.cell(br,2+j),border=BRD,fill=F_ENTRY,nf='#,##0')
    ws.cell(br,SUM_C,f"=SUM({cl(2)}{br}:{cl(2+N_AG-1)}{br})")
    sc(ws.cell(br,SUM_C),BOLD,border=BRD,fill=F_FORM,nf='#,##0')
    allsum = "+".join([f"{cl(SUM_C)}{BLDG_START+k}" for k in range(N_BLDG)])
    ws.cell(br,AW_C,f"=IFERROR({cl(SUM_C)}{br}/({allsum}),0)")
    sc(ws.cell(br,AW_C),BOLD,border=BRD,fill=F_FORM,nf='0.0000')
    ws.cell(br,W_C,f"={cl(AW_C)}{br}"); sc(ws.cell(br,W_C),BOLD,border=BRD,fill=F_ENTRY,nf='0.0000')

chk_br = BLDG_START + N_BLDG + 1
ws.cell(chk_br,1,"✓ ∑ вес корп."); sc(ws.cell(chk_br,1),Font(bold=True,italic=True,size=9),align=LW)
wsum = "+".join([f"{cl(W_C)}{BLDG_START+k}" for k in range(N_BLDG)])
ws.cell(chk_br,W_C,f"={wsum}"); sc(ws.cell(chk_br,W_C),nf='0.0000')
ws.conditional_formatting.add(f"{cl(W_C)}{chk_br}",CellIsRule(operator="notEqual",formula=["1"],fill=F_WARN))
ws.conditional_formatting.add(f"{cl(W_C)}{chk_br}",CellIsRule(operator="equal",formula=["1"],fill=F_OK))

for j,cat in enumerate(ABOVEGROUND):
    r = rref(cat)
    sp = "+".join([f"{cl(2+j)}{BLDG_START+k}" for k in range(N_BLDG)])
    ws.cell(r, 4, f"={sp}")


# ═══════════════════════════════════════════════════════════
# 2. ВВОД
# ═══════════════════════════════════════════════════════════
def build_vvod(title, cats, color):
    wv = wb.create_sheet(title); wv.sheet_properties.tabColor = color
    nc = len(cats); block = nc * 2; vlast = V_DR + N_WEEKS * block - 1
    wv.merge_cells("A1:I1")
    c = wv["A1"]; c.value = f"ВВОД ДАННЫХ — {title}"
    sc(c,B14W,F_TITLE,CTR); sr(wv,1,1,1,9,fill=F_TITLE)
    wv.merge_cells("A2:I2")
    c = wv["A2"]; c.value = "Вводите м². Серые столбцы A-D не редактируйте."
    sc(c,Font(italic=True),F_PAR,CTR)
    for i,h in enumerate(["Дата","Тип","Код","Категория","Пол","Стены","Потолок","Мебл.","✓"],1):
        c = wv.cell(3,i,h); sc(c,BOLD,F_HDR,CTR,BRD)
    for wi in range(N_WEEKS):
        bs = V_DR + wi * block
        for ti,tn in enumerate(["Черновая","Чистовая"]):
            for ci,cat in enumerate(cats):
                r = bs + ti*nc + ci
                if wi==0 and ti==0 and ci==0: wv.cell(r,1,dates[0])
                elif ti==0 and ci==0: wv.cell(r,1,f"=A{V_DR+(wi-1)*block}+7")
                else: wv.cell(r,1,f"=A{bs}")
                sc(wv.cell(r,1),border=BRD,nf="DD.MM.YYYY",fill=F_PREFILL)
                wv.cell(r,2,tn); sc(wv.cell(r,2),border=BRD,fill=F_PREFILL)
                wv.cell(r,3,cat["code"]); sc(wv.cell(r,3),border=BRD,fill=F_PREFILL)
                wv.cell(r,4,cat["name"]); sc(wv.cell(r,4),border=BRD,fill=F_PREFILL,align=LW)
                for ec in range(5,9): sc(wv.cell(r,ec),border=BRD,fill=F_ENTRY,nf='#,##0.00')
                cnt = '=COUNTA(E{0}:H{0})&"/4"' if tn=="Чистовая" and cat.get("has_furn") else '=COUNTA(E{0}:G{0})&"/3"'
                wv.cell(r,9,cnt.format(r)); sc(wv.cell(r,9),border=BRD,fill=F_FORM,align=CTR)
    wv.conditional_formatting.add(f"A{V_DR}:I{vlast}",
        FormulaRule(formula=[f"AND($A{V_DR}<=TODAY(),$A{V_DR}+6>=TODAY())"],
                    fill=PatternFill("solid",fgColor="FFFF99")))
    dv3 = DataValidation(type="decimal",operator="greaterThanOrEqual",formula1=0)
    wv.add_data_validation(dv3); dv3.add(f"E{V_DR}:H{vlast}")
    wv.column_dimensions["A"].width=12; wv.column_dimensions["B"].width=12
    wv.column_dimensions["C"].width=8; wv.column_dimensions["D"].width=38
    for x in "EFGH": wv.column_dimensions[x].width=12
    wv.column_dimensions["I"].width=8; wv.freeze_panes=f"E{V_DR}"
    print(f"  {title}: {vlast} rows ({N_WEEKS}w x {block})")
    return vlast

VL_UG = build_vvod("ВВОД ПОДЗЕМ", UNDERGROUND, "00B050")
VL_AG = None
for bn in BUILDINGS:
    VL_AG = build_vvod(f"ВВОД {bn}", ABOVEGROUND, "00B050")


# ═══════════════════════════════════════════════════════════
# 3. ПЛАН + ФАКТ
# ═══════════════════════════════════════════════════════════
def build_sheet(title, cats, is_fin, color, tot_label="Итого"):
    ws = wb.create_sheet(title); ws.sheet_properties.tabColor = color
    col=2; cmap={}
    for cat in cats:
        s=col; nc=4 if (is_fin and cat.get("has_furn")) else 3
        cols=list(range(col,col+nc)); cmap[cat["code"]]=(s,col+nc-1,cols); col+=nc
    cat_e=col-1; col+=1; tot=[col,col+1,col+2]; col+=3; maxc=col-1
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=cat_e)
    c=ws.cell(1,1,title.upper()); sc(c,B14W,F_TITLE,CTR); sr(ws,1,1,1,cat_e,fill=F_TITLE)
    ws.merge_cells(start_row=1,start_column=tot[0],end_row=1,end_column=tot[2])
    c=ws.cell(1,tot[0],tot_label); sc(c,BW,F_TITLE,CTR); sr(ws,1,1,tot[0],tot[2],fill=F_TITLE)
    ws.merge_cells(start_row=2,start_column=1,end_row=3,end_column=1)
    c=ws.cell(2,1,"Дата"); sc(c,BOLD,F_HDR,CTR,BRD)
    fl = F_UG if cats[0]["code"].startswith("8.1") else F_AG
    ws.merge_cells(start_row=2,start_column=2,end_row=2,end_column=cat_e)
    gl = "Подземная часть" if cats[0]["code"].startswith("8.1") else "Надземная часть"
    c=ws.cell(2,2,gl); sc(c,BOLD,fl,CTR,BRD); sr(ws,2,2,2,cat_e,fill=fl,border=BRD)
    for cat in cats:
        s,e,_=cmap[cat["code"]]
        if s!=e: ws.merge_cells(start_row=3,start_column=s,end_row=3,end_column=e)
        c=ws.cell(3,s,f"{cat['code']} {cat['name']}"); sc(c,BOLD,fl,CTR,BRD); sr(ws,3,3,s,e,fill=fl,border=BRD)
    for i,sf in enumerate(SURFS): c=ws.cell(3,tot[i],sf); sc(c,BOLD,F_TOT,CTR,BRD)
    for cat in cats:
        _,_,cols=cmap[cat["code"]]; sfs=list(SURFS)
        if is_fin and cat.get("has_furn"): sfs.append("Мебл.")
        for i,sf in enumerate(sfs): c=ws.cell(4,cols[i],sf); sc(c,BOLD,fl,CTR,BRD)
    sr(ws,4,4,tot[0],tot[2],fill=F_TOT,border=BRD)
    is_plan = "ПЛАН" in title; cfill = F_PLAN if is_plan else None
    for di in range(N_WEEKS):
        r=DR+di
        if di==0: ws.cell(r,1,dates[0])
        else: ws.cell(r,1,f"=A{r-1}+7")
        sc(ws.cell(r,1),border=BRD,nf="DD.MM.YYYY")
        for cat in cats:
            _,_,cols=cmap[cat["code"]]
            for ci in cols: sc(ws.cell(r,ci),border=BRD,nf='#,##0.00',fill=cfill)
        for si in range(3):
            parts=[f"{cl(cmap[c['code']][2][si])}{r}" for c in cats]
            ws.cell(r,tot[si],f"={'+'.join(parts)}"); sc(ws.cell(r,tot[si]),border=BRD,fill=F_TOT,nf='#,##0.00')
    ws.conditional_formatting.add(f"A{DR}:A{LDR}",
        FormulaRule(formula=[f"AND($A{DR}<=TODAY(),$A{DR}+6>=TODAY())"],
                    fill=PatternFill("solid",fgColor="FFFF99")))
    ws.column_dimensions["A"].width=12
    for ci in range(2,maxc+1): ws.column_dimensions[cl(ci)].width=10
    ws.freeze_panes=ws.cell(DR,2)
    dv=DataValidation(type="decimal",operator="greaterThanOrEqual",formula1=0)
    ws.add_data_validation(dv); dv.add(f"B{DR}:{cl(cat_e)}{LDR}")
    return cmap

ug_prc = build_sheet("ПЛАН ЧЕРН ПОДЗЕМ", UNDERGROUND, False, "4472C4", "Итого подзем.")
ug_pfc = build_sheet("ПЛАН ЧИСТ ПОДЗЕМ", UNDERGROUND, True, "4472C4", "Итого подзем.")
bg_prc={}; bg_pfc={}
for bn in BUILDINGS:
    bg_prc[bn] = build_sheet(f"ПЛАН ЧЕРН {bn}", ABOVEGROUND, False, "4472C4", f"Итого {bn}")
    bg_pfc[bn] = build_sheet(f"ПЛАН ЧИСТ {bn}", ABOVEGROUND, True, "4472C4", f"Итого {bn}")

ug_rc = build_sheet("ФАКТ ЧЕРН ПОДЗЕМ", UNDERGROUND, False, "BF8F00", "Итого подзем.")
ug_fc = build_sheet("ФАКТ ЧИСТ ПОДЗЕМ", UNDERGROUND, True, "548235", "Итого подзем.")
bg_rc={}; bg_fc={}
for bn in BUILDINGS:
    bg_rc[bn] = build_sheet(f"ФАКТ ЧЕРН {bn}", ABOVEGROUND, False, "BF8F00", f"Итого {bn}")
    bg_fc[bn] = build_sheet(f"ФАКТ ЧИСТ {bn}", ABOVEGROUND, True, "548235", f"Итого {bn}")


# ═══════════════════════════════════════════════════════════
# 4. ОТЧЕТ
# ═══════════════════════════════════════════════════════════
ws_rep = wb.create_sheet("ОТЧЕТ"); ws_rep.sheet_properties.tabColor = "7030A0"
ws_rep.merge_cells("A1:I1")
c=ws_rep["A1"]; c.value="ОТЧЁТ ПО ОТДЕЛОЧНЫМ РАБОТАМ"
sc(c,B14W,F_TITLE,CTR); sr(ws_rep,1,1,1,9,fill=F_TITLE)
ws_rep.merge_cells("A2:I2")
c=ws_rep["A2"]; c.value="⚠ Автоматический. Не вносите данные напрямую."
sc(c,Font(bold=True,italic=True,color="FF0000"),F_WARN,CTR)

rr=4

def write_section(ws, start_r, title, cats, rcm, fcm, rsh, fsh, vsh, vl, plan_fn=None):
    r=start_r; main_rows={}
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9)
    c=ws.cell(r,1,title); sc(c,B12W,F_TITLE,CTR); sr(ws,r,r,1,9,fill=F_TITLE); r+=1
    for i,h in enumerate(["Код","Название","Тип","Вес ур.4","Вес ур.5","План м²","Факт м²","Факт %","Взвеш. %"],1):
        c=ws.cell(r,i,h); sc(c,BOLD,F_HDR,CTR,BRD)
    r+=1
    rn=f"'{rsh}'"; fn=f"'{fsh}'"
    for cat in cats:
        rr_=rref(cat); code=cat["code"]; mr=r; main_rows[code]=mr
        ws.cell(r,1,code); sc(ws.cell(r,1),B11,border=BRD)
        ws.cell(r,2,f"={REF}!C{rr_}"); sc(ws.cell(r,2),B11,border=BRD,align=LW)
        ws.cell(r,3,"Главная"); sc(ws.cell(r,3),border=BRD)
        sc(ws.cell(r,4),border=BRD); sc(ws.cell(r,5),border=BRD)
        ws.cell(r,6,plan_fn(cat) if plan_fn else f"={REF}!D{rr_}")
        sc(ws.cell(r,6),border=BRD,fill=F_FORM,nf='#,##0')
        sc(ws.cell(r,7),border=BRD,fill=F_FORM,nf='#,##0.00')
        ws.cell(r,8,f"=IF(F{r}=0,0,G{r}/F{r})"); sc(ws.cell(r,8),border=BRD,fill=F_FORM,nf='0.00%')
        sc(ws.cell(r,9),border=BRD,fill=F_FORM,nf='0.00%'); r+=1
        phase_rows=[]
        for phase,pname,fsn,wc,cm in [("r","Черновая",rn,"E",rcm),("f","Чистовая",fn,"F",fcm)]:
            tname="Черновая" if phase=="r" else "Чистовая"
            pr=r; phase_rows.append(pr)
            ws.cell(r,2,pname); sc(ws.cell(r,2),BOLD,border=BRD,align=LW)
            sc(ws.cell(r,3),border=BRD)
            ws.cell(r,4,f"={REF}!{wc}{rr_}"); sc(ws.cell(r,4),border=BRD,fill=F_FORM,nf='0.00')
            sc(ws.cell(r,5),border=BRD)
            ws.cell(r,6,f"=F{mr}"); sc(ws.cell(r,6),border=BRD,fill=F_FORM,nf='#,##0')
            sc(ws.cell(r,7),border=BRD,fill=F_FORM,nf='#,##0.00')
            ws.cell(r,8,f"=IF(F{r}=0,0,G{r}/F{r})"); sc(ws.cell(r,8),border=BRD,fill=F_FORM,nf='0.00%')
            ws.cell(r,9,f"=H{r}*D{r}"); sc(ws.cell(r,9),border=BRD,fill=F_FORM,nf='0.00%')
            sc(ws.cell(r,1),border=BRD); r+=1
            _,_,scols=cm[code]; srows=[]; ns=len(scols)
            sn=list(SURFS)
            if ns==4: sn.append("Мебл.")
            for si in range(ns):
                ws.cell(r,2,f"   {sn[si]}"); sc(ws.cell(r,2),border=BRD,align=LW)
                ws.cell(r,3,"Вид"); sc(ws.cell(r,3),border=BRD)
                sc(ws.cell(r,4),border=BRD)
                sw=cl(8+si) if phase=="r" else cl(12+si)
                ws.cell(r,5,f"={REF}!{sw}{rr_}"); sc(ws.cell(r,5),border=BRD,fill=F_FORM,nf='0.00')
                ws.cell(r,6,f"=F{mr}"); sc(ws.cell(r,6),border=BRD,fill=F_FORM,nf='#,##0')
                fc_=cl(scols[si])
                fp=f"SUM({fsn}!{fc_}{DR}:{fc_}{LDR})"
                vp=vsum(sn[si],code,tname,vsh,vl)
                ws.cell(r,7,f"={fp}+{vp}"); sc(ws.cell(r,7),border=BRD,fill=F_FORM,nf='#,##0.00')
                ws.cell(r,8,f"=IF(F{r}=0,0,G{r}/F{r})"); sc(ws.cell(r,8),border=BRD,fill=F_FORM,nf='0.00%')
                ws.cell(r,9,f"=H{r}*E{r}"); sc(ws.cell(r,9),border=BRD,fill=F_FORM,nf='0.00%')
                sc(ws.cell(r,1),border=BRD); srows.append(r); r+=1
            wp=[f"G{s}*E{s}" for s in srows]; ws.cell(pr,7,f"={'+'.join(wp)}")
        ws.cell(mr,7,f"={'+'.join([f'G{p}*D{p}' for p in phase_rows])}")
        ws.cell(mr,9,f"={'+'.join([f'I{p}' for p in phase_rows])}")
        ws.cell(r,2,"✓ Сумма весов ур.4"); sc(ws.cell(r,2),Font(bold=True,italic=True,size=9),align=LW)
        ws.cell(r,4,f"={'+'.join([f'D{p}' for p in phase_rows])}"); sc(ws.cell(r,4),nf='0.00')
        ws.conditional_formatting.add(f"D{r}",CellIsRule(operator="notEqual",formula=["1"],fill=F_WARN))
        ws.conditional_formatting.add(f"D{r}",CellIsRule(operator="equal",formula=["1"],fill=F_OK))
        r+=2
    return r, main_rows

rr, ug_main = write_section(ws_rep, rr, "ПОДЗЕМНАЯ ЧАСТЬ", UNDERGROUND,
    ug_rc, ug_fc, "ФАКТ ЧЕРН ПОДЗЕМ", "ФАКТ ЧИСТ ПОДЗЕМ", "ВВОД ПОДЗЕМ", VL_UG)

bg_main = {}
for bn in BUILDINGS:
    def pf(cat, b=bn):
        j = ABOVEGROUND.index(cat)
        return f"={REF}!{cl(2+j)}{BLDG_ROWS[b]}"
    rr, bm = write_section(ws_rep, rr, f"НАДЗЕМНАЯ ЧАСТЬ — {bn}", ABOVEGROUND,
        bg_rc[bn], bg_fc[bn], f"ФАКТ ЧЕРН {bn}", f"ФАКТ ЧИСТ {bn}", f"ВВОД {bn}", VL_AG, plan_fn=pf)
    bg_main[bn] = bm

ws_rep.column_dimensions["A"].width=8; ws_rep.column_dimensions["B"].width=52; ws_rep.column_dimensions["C"].width=12
for x in "DEFGHI": ws_rep.column_dimensions[x].width=14
ws_rep.freeze_panes="A4"

# ── Level 2 & 1 aggregation ──
rr += 1
ws_rep.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=9)
c = ws_rep.cell(rr,1,"АГРЕГАЦИЯ: УРОВЕНЬ 3 → УРОВЕНЬ 2 → УРОВЕНЬ 1")
sc(c,B12W,F_TITLE,CTR); sr(ws_rep,rr,rr,1,9,fill=F_TITLE); rr+=1
for i,h in enumerate(["","Название","","Вес","","План м²","","Факт %","Взвеш. %"],1):
    if h: c=ws_rep.cell(rr,i,h); sc(c,BOLD,F_HDR,CTR,BRD)
    else: sc(ws_rep.cell(rr,i),fill=F_HDR,border=BRD)
rr+=1

lvl2={}
# 8.1
lvl2["8.1"]=rr
ws_rep.cell(rr,1,"8.1"); sc(ws_rep.cell(rr,1),font=B11,fill=F_LVL,border=BRD)
ws_rep.cell(rr,2,"Подземная часть"); sc(ws_rep.cell(rr,2),font=B11,fill=F_LVL,align=LW,border=BRD)
sc(ws_rep.cell(rr,3),fill=F_LVL,border=BRD)
ws_rep.cell(rr,4,f"={REF}!V{SUM_UG}"); sc(ws_rep.cell(rr,4),font=BOLD,fill=F_LVL,border=BRD,nf='0.0000')
sc(ws_rep.cell(rr,5),fill=F_LVL,border=BRD)
pp=[f"F{ug_main[c['code']]}" for c in UNDERGROUND]
ws_rep.cell(rr,6,f"={'+'.join(pp)}"); sc(ws_rep.cell(rr,6),fill=F_LVL,border=BRD,nf='#,##0')
sc(ws_rep.cell(rr,7),fill=F_LVL,border=BRD)
fp=[f"{REF}!V{rref(c)}*H{ug_main[c['code']]}" for c in UNDERGROUND]
ws_rep.cell(rr,8,f"={'+'.join(fp)}"); sc(ws_rep.cell(rr,8),font=BOLD,fill=F_LVL,border=BRD,nf='0.00%')
ws_rep.cell(rr,9,f"=H{rr}*D{rr}"); sc(ws_rep.cell(rr,9),font=BOLD,fill=F_LVL,border=BRD,nf='0.00%')
rr+=1

# Per-building 8.2
bl2={}
for bn in BUILDINGS:
    bl2[bn]=rr
    ws_rep.cell(rr,1,"8.2"); sc(ws_rep.cell(rr,1),font=B11,fill=F_BLDG,border=BRD)
    ws_rep.cell(rr,2,f"Надземная {bn}"); sc(ws_rep.cell(rr,2),font=B11,fill=F_BLDG,align=LW,border=BRD)
    sc(ws_rep.cell(rr,3),fill=F_BLDG,border=BRD)
    ws_rep.cell(rr,4,f"={REF}!{cl(W_C)}{BLDG_ROWS[bn]}")
    sc(ws_rep.cell(rr,4),font=BOLD,fill=F_BLDG,border=BRD,nf='0.0000')
    sc(ws_rep.cell(rr,5),fill=F_BLDG,border=BRD)
    pp=[f"F{bg_main[bn][c['code']]}" for c in ABOVEGROUND]
    ws_rep.cell(rr,6,f"={'+'.join(pp)}"); sc(ws_rep.cell(rr,6),fill=F_BLDG,border=BRD,nf='#,##0')
    sc(ws_rep.cell(rr,7),fill=F_BLDG,border=BRD)
    fp=[f"{REF}!V{rref(c)}*H{bg_main[bn][c['code']]}" for c in ABOVEGROUND]
    ws_rep.cell(rr,8,f"={'+'.join(fp)}"); sc(ws_rep.cell(rr,8),font=BOLD,fill=F_BLDG,border=BRD,nf='0.00%')
    ws_rep.cell(rr,9,f"=H{rr}*D{rr}"); sc(ws_rep.cell(rr,9),font=BOLD,fill=F_BLDG,border=BRD,nf='0.00%')
    rr+=1

# 8.2 total
lvl2["8.2"]=rr
ws_rep.cell(rr,1,"8.2"); sc(ws_rep.cell(rr,1),font=B11,fill=F_LVL,border=BRD)
ws_rep.cell(rr,2,"Надземная часть (все корпуса)"); sc(ws_rep.cell(rr,2),font=B11,fill=F_LVL,align=LW,border=BRD)
sc(ws_rep.cell(rr,3),fill=F_LVL,border=BRD)
ws_rep.cell(rr,4,f"={REF}!V{SUM_AG}"); sc(ws_rep.cell(rr,4),font=BOLD,fill=F_LVL,border=BRD,nf='0.0000')
sc(ws_rep.cell(rr,5),fill=F_LVL,border=BRD)
pp=[f"F{bl2[b]}" for b in BUILDINGS]
ws_rep.cell(rr,6,f"={'+'.join(pp)}"); sc(ws_rep.cell(rr,6),fill=F_LVL,border=BRD,nf='#,##0')
sc(ws_rep.cell(rr,7),fill=F_LVL,border=BRD)
fp=[f"D{bl2[b]}*H{bl2[b]}" for b in BUILDINGS]
ws_rep.cell(rr,8,f"={'+'.join(fp)}"); sc(ws_rep.cell(rr,8),font=BOLD,fill=F_LVL,border=BRD,nf='0.00%')
ws_rep.cell(rr,9,f"=H{rr}*D{rr}"); sc(ws_rep.cell(rr,9),font=BOLD,fill=F_LVL,border=BRD,nf='0.00%')
rr+=1

# Checks
ws_rep.cell(rr,2,"✓ ∑ вес ур.2"); sc(ws_rep.cell(rr,2),Font(bold=True,italic=True,size=9),align=LW)
ws_rep.cell(rr,4,f"=D{lvl2['8.1']}+D{lvl2['8.2']}"); sc(ws_rep.cell(rr,4),nf='0.0000')
ws_rep.conditional_formatting.add(f"D{rr}",CellIsRule(operator="notEqual",formula=["1"],fill=F_WARN))
ws_rep.conditional_formatting.add(f"D{rr}",CellIsRule(operator="equal",formula=["1"],fill=F_OK))
rr+=1
ws_rep.cell(rr,2,"✓ ∑ вес корп."); sc(ws_rep.cell(rr,2),Font(bold=True,italic=True,size=9),align=LW)
bwp="+".join([f"D{bl2[b]}" for b in BUILDINGS])
ws_rep.cell(rr,4,f"={bwp}"); sc(ws_rep.cell(rr,4),nf='0.0000')
ws_rep.conditional_formatting.add(f"D{rr}",CellIsRule(operator="notEqual",formula=[f"D{lvl2['8.2']}"],fill=F_WARN))
ws_rep.conditional_formatting.add(f"D{rr}",CellIsRule(operator="equal",formula=[f"D{lvl2['8.2']}"],fill=F_OK))
rr+=2

# Level 1
ws_rep.cell(rr,1,"8"); sc(ws_rep.cell(rr,1),font=B11,fill=F_LVL1,border=BRD)
ws_rep.cell(rr,2,"ОТДЕЛКА — ОБЩИЙ %"); sc(ws_rep.cell(rr,2),font=Font(bold=True,size=12),fill=F_LVL1,align=LW,border=BRD)
for c_ in [3,4,5]: sc(ws_rep.cell(rr,c_),fill=F_LVL1,border=BRD)
ws_rep.cell(rr,6,f"=F{lvl2['8.1']}+F{lvl2['8.2']}")
sc(ws_rep.cell(rr,6),font=Font(bold=True,size=12),fill=F_LVL1,border=BRD,nf='#,##0')
sc(ws_rep.cell(rr,7),fill=F_LVL1,border=BRD)
ws_rep.cell(rr,8,f"=I{lvl2['8.1']}+I{lvl2['8.2']}")
sc(ws_rep.cell(rr,8),font=Font(bold=True,size=14),fill=F_LVL1,border=BRD,nf='0.00%')
sc(ws_rep.cell(rr,9),fill=F_LVL1,border=BRD)


# ═══════════════════════════════════════════════════════════
# 5. SIGNAL — горизонтальная раскладка, агрегатные карточки
# ═══════════════════════════════════════════════════════════
ws_sig = wb.create_sheet("SIGNAL"); ws_sig.sheet_properties.tabColor = "4472C4"
ws_sig.cell(1,1,"Дата отчета"); sc(ws_sig.cell(1,1),BOLD,F_PAR,LW,BRD)
ws_sig.cell(1,2,datetime(2026,6,29)); sc(ws_sig.cell(1,2),BOLD,F_PAR,border=BRD,nf="DD.MM.YYYY")
ws_sig.cell(1,2).comment=Comment("Замените на =TODAY()","Шаблон")

# Each group = 3 cards (total, rough, finish) × 2 cols (Plan, Fact) = 6 cols
GRP_W = 6
grps = [{"label":"8.1 Подземная часть","pa":f"={REF}!D{SUM_UG}"}]
for bn in BUILDINGS:
    psum = "+".join([f"{REF}!{cl(2+j)}{BLDG_ROWS[bn]}" for j in range(N_AG)])
    grps.append({"label":f"8.2 Надземная {bn}","pa":f"={psum}"})
grps.append({"label":"8.2 Надзем. (все)","pa":f"={REF}!D{SUM_AG}"})
grps.append({"label":"8. Отделка общий","pa":f"={REF}!D{SUM_UG}+{REF}!D{SUM_AG}"})
N_GRP = len(grps)
CTYPES = ["Общий","Черновая","Чистовая"]

def sc_(gi, ci, pf):
    return 2 + gi*GRP_W + ci*2 + pf

# Row 3: group headers (merged across 6 cols)
for gi, g in enumerate(grps):
    gc = 2 + gi*GRP_W
    ws_sig.merge_cells(start_row=3, start_column=gc, end_row=3, end_column=gc+GRP_W-1)
    c = ws_sig.cell(3, gc, g["label"])
    sc(c, BW, F_TITLE, CTR, BRD); sr(ws_sig, 3, 3, gc, gc+GRP_W-1, fill=F_TITLE, border=BRD)

# Row 4: card sub-headers (Общий / Черновая / Чистовая, merged 2 cols each)
for gi in range(N_GRP):
    for ci, ct in enumerate(CTYPES):
        cc = sc_(gi, ci, 0)
        ws_sig.merge_cells(start_row=4, start_column=cc, end_row=4, end_column=cc+1)
        c = ws_sig.cell(4, cc, ct); sc(c, BOLD, F_HDR, CTR, BRD)
        sc(ws_sig.cell(4, cc+1), fill=F_HDR, border=BRD)

# Rows 5-16: config fields (names in col A, values across card columns)
CFG_R = 5
SIG_CFG = [("Заголовок","План-факт по объемам"),("Статус",True),("Url изображения",""),
           ("Тип","planFact2"),("",""),("Дата","=$B$1"),("Гистограмма",False),
           ("По месяцам",False),("Всего",None),("Тип",None),("Ед. изм.","м2"),
           ("Учитывать дату карточки",False)]

for fi, (fld, fv) in enumerate(SIG_CFG):
    r = CFG_R + fi
    if fld:
        ws_sig.cell(r, 1, fld); sc(ws_sig.cell(r, 1), BOLD, F_CFG, LW, BRD)
    for gi, g in enumerate(grps):
        for ci, ct in enumerate(CTYPES):
            cc = sc_(gi, ci, 0)
            if fld == "Всего":
                ws_sig.cell(r, cc, g["pa"])
            elif fld == "Тип" and fi == 9:
                ws_sig.cell(r, cc, f"{g['label']}: {ct}")
            elif fld == "Дата":
                ws_sig.cell(r, cc, "=$B$1")
                sc(ws_sig.cell(r, cc), nf="DD.MM.YYYY")
            elif fld:
                ws_sig.cell(r, cc, fv)
            sc(ws_sig.cell(r, cc), border=BRD)
            sc(ws_sig.cell(r, cc+1), border=BRD)
    if fld in CFG_COMMENTS:
        ws_sig.cell(r, 1).comment = Comment(CFG_COMMENTS[fld], "SIGNAL")

# Row 17: data headers
HDR_R = CFG_R + len(SIG_CFG)
ws_sig.cell(HDR_R, 1, "Дата"); sc(ws_sig.cell(HDR_R, 1), BOLD, F_HDR, CTR, BRD)
for gi in range(N_GRP):
    for ci in range(3):
        pc = sc_(gi, ci, 0); fc = sc_(gi, ci, 1)
        ws_sig.cell(HDR_R, pc, "План"); sc(ws_sig.cell(HDR_R, pc), BOLD, F_HDR, CTR, BRD)
        ws_sig.cell(HDR_R, fc, "Факт"); sc(ws_sig.cell(HDR_R, fc), BOLD, F_HDR, CTR, BRD)

# Rows 18+: weekly data
DATA_R = HDR_R + 1
gi_82 = 1 + N_BLDG
gi_all = 2 + N_BLDG

for di in range(N_WEEKS):
    r = DATA_R + di
    fr = DR + di
    dc = f"$A{r}"
    if di == 0:
        ws_sig.cell(r, 1, dates[0])
    else:
        ws_sig.cell(r, 1, f"=A{r-1}+7")
    sc(ws_sig.cell(r, 1), border=BRD, nf="DD.MM.YYYY")

    # 8.1 underground rough
    rp = []; rf = []
    for cat in UNDERGROUND:
        rr_ = rref(cat); _, _, prsc = ug_prc[cat["code"]]; _, _, rsc = ug_rc[cat["code"]]
        for si in range(3):
            rp.append(f"'ПЛАН ЧЕРН ПОДЗЕМ'!{cl(prsc[si])}{fr}*{REF}!${cl(8+si)}${rr_}")
            vv = vsum(SURFS[si], cat["code"], "Черновая", "ВВОД ПОДЗЕМ", VL_UG, dc)
            rf.append(f"('ФАКТ ЧЕРН ПОДЗЕМ'!{cl(rsc[si])}{fr}+{vv})*{REF}!${cl(8+si)}${rr_}")
    ws_sig.cell(r, sc_(0,1,0), "="+"+".join(rp))
    sc(ws_sig.cell(r, sc_(0,1,0)), border=BRD, fill=F_PLAN, nf='#,##0.00')
    ws_sig.cell(r, sc_(0,1,1), "="+"+".join(rf))
    sc(ws_sig.cell(r, sc_(0,1,1)), border=BRD, fill=F_FORM, nf='#,##0.00')

    # 8.1 underground finish
    fp = []; ff = []
    for cat in UNDERGROUND:
        rr_ = rref(cat); _, _, pfsc = ug_pfc[cat["code"]]; _, _, fsc = ug_fc[cat["code"]]
        sn = list(SURFS)
        if len(pfsc) == 4: sn.append("Мебл.")
        for si in range(len(pfsc)):
            fp.append(f"'ПЛАН ЧИСТ ПОДЗЕМ'!{cl(pfsc[si])}{fr}*{REF}!${cl(12+si)}${rr_}")
            vv = vsum(sn[si], cat["code"], "Чистовая", "ВВОД ПОДЗЕМ", VL_UG, dc)
            ff.append(f"('ФАКТ ЧИСТ ПОДЗЕМ'!{cl(fsc[si])}{fr}+{vv})*{REF}!${cl(12+si)}${rr_}")
    ws_sig.cell(r, sc_(0,2,0), "="+"+".join(fp))
    sc(ws_sig.cell(r, sc_(0,2,0)), border=BRD, fill=F_PLAN, nf='#,##0.00')
    ws_sig.cell(r, sc_(0,2,1), "="+"+".join(ff))
    sc(ws_sig.cell(r, sc_(0,2,1)), border=BRD, fill=F_FORM, nf='#,##0.00')

    # 8.1 total = rough + finish
    ws_sig.cell(r, sc_(0,0,0), f"={cl(sc_(0,1,0))}{r}+{cl(sc_(0,2,0))}{r}")
    sc(ws_sig.cell(r, sc_(0,0,0)), border=BRD, fill=F_PLAN, nf='#,##0.00')
    ws_sig.cell(r, sc_(0,0,1), f"={cl(sc_(0,1,1))}{r}+{cl(sc_(0,2,1))}{r}")
    sc(ws_sig.cell(r, sc_(0,0,1)), border=BRD, fill=F_FORM, nf='#,##0.00')

    # Per-building 8.2
    for bi, bn in enumerate(BUILDINGS):
        gi = 1 + bi
        prsh = f"ПЛАН ЧЕРН {bn}"; pfsh = f"ПЛАН ЧИСТ {bn}"
        frsh = f"ФАКТ ЧЕРН {bn}"; ffsh = f"ФАКТ ЧИСТ {bn}"
        vsh = f"ВВОД {bn}"
        rp = []; rf = []
        for cat in ABOVEGROUND:
            rr_ = rref(cat)
            _, _, prsc = bg_prc[bn][cat["code"]]; _, _, rsc = bg_rc[bn][cat["code"]]
            for si in range(3):
                rp.append(f"'{prsh}'!{cl(prsc[si])}{fr}*{REF}!${cl(8+si)}${rr_}")
                vv = vsum(SURFS[si], cat["code"], "Черновая", vsh, VL_AG, dc)
                rf.append(f"('{frsh}'!{cl(rsc[si])}{fr}+{vv})*{REF}!${cl(8+si)}${rr_}")
        ws_sig.cell(r, sc_(gi,1,0), "="+"+".join(rp))
        sc(ws_sig.cell(r, sc_(gi,1,0)), border=BRD, fill=F_PLAN, nf='#,##0.00')
        ws_sig.cell(r, sc_(gi,1,1), "="+"+".join(rf))
        sc(ws_sig.cell(r, sc_(gi,1,1)), border=BRD, fill=F_FORM, nf='#,##0.00')
        fp = []; ff = []
        for cat in ABOVEGROUND:
            rr_ = rref(cat)
            _, _, pfsc = bg_pfc[bn][cat["code"]]; _, _, fsc = bg_fc[bn][cat["code"]]
            sn = list(SURFS)
            if len(pfsc) == 4: sn.append("Мебл.")
            for si in range(len(pfsc)):
                fp.append(f"'{pfsh}'!{cl(pfsc[si])}{fr}*{REF}!${cl(12+si)}${rr_}")
                vv = vsum(sn[si], cat["code"], "Чистовая", vsh, VL_AG, dc)
                ff.append(f"('{ffsh}'!{cl(fsc[si])}{fr}+{vv})*{REF}!${cl(12+si)}${rr_}")
        ws_sig.cell(r, sc_(gi,2,0), "="+"+".join(fp))
        sc(ws_sig.cell(r, sc_(gi,2,0)), border=BRD, fill=F_PLAN, nf='#,##0.00')
        ws_sig.cell(r, sc_(gi,2,1), "="+"+".join(ff))
        sc(ws_sig.cell(r, sc_(gi,2,1)), border=BRD, fill=F_FORM, nf='#,##0.00')
        ws_sig.cell(r, sc_(gi,0,0), f"={cl(sc_(gi,1,0))}{r}+{cl(sc_(gi,2,0))}{r}")
        sc(ws_sig.cell(r, sc_(gi,0,0)), border=BRD, fill=F_PLAN, nf='#,##0.00')
        ws_sig.cell(r, sc_(gi,0,1), f"={cl(sc_(gi,1,1))}{r}+{cl(sc_(gi,2,1))}{r}")
        sc(ws_sig.cell(r, sc_(gi,0,1)), border=BRD, fill=F_FORM, nf='#,##0.00')

    # 8.2 total = sum of buildings
    for ci in range(3):
        pp = "+".join([f"{cl(sc_(1+bi,ci,0))}{r}" for bi in range(N_BLDG)])
        ws_sig.cell(r, sc_(gi_82,ci,0), f"={pp}")
        sc(ws_sig.cell(r, sc_(gi_82,ci,0)), border=BRD, fill=F_PLAN, nf='#,##0.00')
        fp_ = "+".join([f"{cl(sc_(1+bi,ci,1))}{r}" for bi in range(N_BLDG)])
        ws_sig.cell(r, sc_(gi_82,ci,1), f"={fp_}")
        sc(ws_sig.cell(r, sc_(gi_82,ci,1)), border=BRD, fill=F_FORM, nf='#,##0.00')

    # 8. overall = 8.1 + 8.2
    for ci in range(3):
        ws_sig.cell(r, sc_(gi_all,ci,0), f"={cl(sc_(0,ci,0))}{r}+{cl(sc_(gi_82,ci,0))}{r}")
        sc(ws_sig.cell(r, sc_(gi_all,ci,0)), border=BRD, fill=F_PLAN, nf='#,##0.00')
        ws_sig.cell(r, sc_(gi_all,ci,1), f"={cl(sc_(0,ci,1))}{r}+{cl(sc_(gi_82,ci,1))}{r}")
        sc(ws_sig.cell(r, sc_(gi_all,ci,1)), border=BRD, fill=F_FORM, nf='#,##0.00')

last_col = sc_(N_GRP-1, 2, 1)
ws_sig.column_dimensions["A"].width = 22
for ci in range(2, last_col+1):
    ws_sig.column_dimensions[cl(ci)].width = 12
ws_sig.freeze_panes = ws_sig.cell(DATA_R, 2)
ws_sig.conditional_formatting.add(f"A{DATA_R}:A{DATA_R+N_WEEKS-1}",
    FormulaRule(formula=[f"AND($A{DATA_R}<=TODAY(),$A{DATA_R}+6>=TODAY())"],
               fill=PatternFill("solid",fgColor="FFFF99")))
print(f"  SIGNAL: {DATA_R+N_WEEKS-1}r x {last_col}c ({N_GRP} groups x 3 cards)")


# ═══════════════════════════════════════════════════════════
# Reorder + Save
# ═══════════════════════════════════════════════════════════
order = ["СПРАВОЧНИК","ВВОД ПОДЗЕМ"]
for b in BUILDINGS: order.append(f"ВВОД {b}")
order += ["ПЛАН ЧЕРН ПОДЗЕМ","ПЛАН ЧИСТ ПОДЗЕМ"]
for b in BUILDINGS: order += [f"ПЛАН ЧЕРН {b}",f"ПЛАН ЧИСТ {b}"]
order += ["ФАКТ ЧЕРН ПОДЗЕМ","ФАКТ ЧИСТ ПОДЗЕМ"]
for b in BUILDINGS: order += [f"ФАКТ ЧЕРН {b}",f"ФАКТ ЧИСТ {b}"]
order += ["ОТЧЕТ","SIGNAL"]
for i,name in enumerate(order):
    idx = wb.sheetnames.index(name); wb.move_sheet(name, offset=i-idx)

out = r"\\mr.ru\Service\Personal\ignatov_i\Documents\CloudCode\Отделка SIGNAL\02_Финальная версия - Отделка для SIGNAL v7.xlsx"
wb.save(out); wb.close()
print(f"OK: {out}")
wb2 = openpyxl.load_workbook(out)
for n in wb2.sheetnames:
    ws=wb2[n]; print(f"  {n}: {ws.max_row}r x {ws.max_column}c")
wb2.close()
