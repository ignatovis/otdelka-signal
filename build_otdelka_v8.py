#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Otdelka v7 → v8:
1. Named ranges for all data tables
2. Remove ВВОД references from SIGNAL and ОТЧЕТ formulas (keep only ФАКТ)
3. Synchronize start dates to single cell СПРАВОЧНИК!B35
"""

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = 'Отделка SIGNAL/02_Финальная версия - Отделка для SIGNAL v7 (для использования в отчете).xlsx'
DST = 'Отделка SIGNAL/02_Финальная версия - Отделка для SIGNAL v8.xlsx'

print(f'Loading: {SRC}')
wb = openpyxl.load_workbook(SRC)
print(f'Sheets: {len(wb.sheetnames)}')


# ────────────────────────────────────────────
# TASK 3: Single start date
# ────────────────────────────────────────────
print('\n── TASK 3: Sync start dates ──')

ws_ref = wb['СПРАВОЧНИК']
ws_ref['A35'] = 'Дата начала проекта'
ws_ref['B35'] = datetime(2026, 6, 1)
print('  СПРАВОЧНИК!B35 = 2026-06-01 (master date)')

START_REF = '=СПРАВОЧНИК!$B$35'

# ВВОД sheets: A4
for sn in ['ВВОД ПОДЗЕМ', 'ВВОД К1', 'ВВОД К2', 'ВВОД К3']:
    wb[sn]['A4'] = START_REF
    print(f'  {sn}!A4 → ref')

# ПЛАН + ФАКТ sheets: A5
for kind in ['ПЛАН', 'ФАКТ']:
    for typ in ['ЧЕРН', 'ЧИСТ']:
        for part in ['ПОДЗЕМ', 'К1', 'К2', 'К3']:
            sn = f'{kind} {typ} {part}'
            wb[sn]['A5'] = START_REF
            print(f'  {sn}!A5 → ref')

# SIGNAL: all hardcoded datetime values in row 14
ws_sig = wb['SIGNAL']
sig_dates_fixed = 0
for c in range(1, ws_sig.max_column + 1):
    v = ws_sig.cell(14, c).value
    if isinstance(v, datetime):
        ws_sig.cell(14, c).value = START_REF
        sig_dates_fixed += 1
print(f'  SIGNAL row 14: {sig_dates_fixed} date cells → ref')


# ────────────────────────────────────────────
# TASK 2: Remove ВВОД from formulas
# ────────────────────────────────────────────
print('\n── TASK 2: Remove ВВОД refs from formulas ──')


def find_closing_paren(s, open_idx):
    """Find matching ) for ( at open_idx."""
    depth = 0
    for i in range(open_idx, len(s)):
        if s[i] == '(':
            depth += 1
        elif s[i] == ')':
            depth -= 1
            if depth == 0:
                return i
    return -1


def remove_sumifs_vvod(formula):
    """Remove all SUMIFS(...ВВОД...) calls and their preceding + from formula."""
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula, False
    if 'ВВОД' not in formula:
        return formula, False

    result = formula
    modified = False

    for _ in range(50):
        pos = -1
        search_from = 0
        while search_from < len(result):
            idx = result.find('SUMIFS(', search_from)
            if idx == -1:
                break
            paren_open = idx + 6
            paren_close = find_closing_paren(result, paren_open)
            if paren_close == -1:
                search_from = idx + 1
                continue
            chunk = result[idx:paren_close + 1]
            if 'ВВОД' in chunk:
                pos = idx
                break
            search_from = paren_close + 1

        if pos == -1:
            break

        paren_open = pos + 6
        paren_close = find_closing_paren(result, paren_open)

        cut_start = pos
        p = pos - 1
        while p >= 0 and result[p] == ' ':
            p -= 1
        if p >= 0 and result[p] == '+':
            cut_start = p

        result = result[:cut_start] + result[paren_close + 1:]
        modified = True

    return result, modified


def process_sheet_formulas(ws, sheet_name):
    """Remove ВВОД SUMIFS from all formulas on a sheet."""
    count = 0
    samples = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if val and isinstance(val, str) and 'ВВОД' in val and val.startswith('='):
                new_val, was_mod = remove_sumifs_vvod(val)
                if was_mod:
                    cell.value = new_val
                    count += 1
                    if len(samples) < 2:
                        col_l = get_column_letter(c)
                        samples.append(f'    {col_l}{r}: ...{new_val[:80]}...')
    print(f'  {sheet_name}: {count} formulas modified')
    for s in samples:
        print(s)
    return count


process_sheet_formulas(wb['SIGNAL'], 'SIGNAL')
process_sheet_formulas(wb['ОТЧЕТ'], 'ОТЧЕТ')


# ────────────────────────────────────────────
# TASK 1: Named ranges
# ────────────────────────────────────────────
print('\n── TASK 1: Named ranges ──')


def add_range(name, sheet_title, range_str):
    ref = f"'{sheet_title}'!{range_str}"
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)
    print(f'  + {name}')


# СПРАВОЧНИК
add_range('Справочник_Категории', 'СПРАВОЧНИК', '$A$2:$W$15')
add_range('Справочник_Корпуса', 'СПРАВОЧНИК', '$A$27:$K$30')
add_range('Справочник_Агрегация', 'СПРАВОЧНИК', '$A$17:$W$19')

# ВВОД sheets
for part, label in [('ПОДЗЕМ', 'Подзем'), ('К1', 'К1'), ('К2', 'К2'), ('К3', 'К3')]:
    sn = f'ВВОД {part}'
    ws = wb[sn]
    mc = get_column_letter(ws.max_column)
    add_range(f'Ввод_{label}', sn, f'$A$3:${mc}${ws.max_row}')

# ПЛАН sheets
for typ, tl in [('ЧЕРН', 'Черн'), ('ЧИСТ', 'Чист')]:
    for part, pl in [('ПОДЗЕМ', 'Подзем'), ('К1', 'К1'), ('К2', 'К2'), ('К3', 'К3')]:
        sn = f'ПЛАН {typ} {part}'
        ws = wb[sn]
        mc = get_column_letter(ws.max_column)
        add_range(f'План_{tl}_{pl}', sn, f'$A$3:${mc}${ws.max_row}')

# ФАКТ sheets
for typ, tl in [('ЧЕРН', 'Черн'), ('ЧИСТ', 'Чист')]:
    for part, pl in [('ПОДЗЕМ', 'Подзем'), ('К1', 'К1'), ('К2', 'К2'), ('К3', 'К3')]:
        sn = f'ФАКТ {typ} {part}'
        ws = wb[sn]
        mc = get_column_letter(ws.max_column)
        add_range(f'Факт_{tl}_{pl}', sn, f'$A$3:${mc}${ws.max_row}')

# ОТЧЕТ
add_range('Отчет_Данные', 'ОТЧЕТ', '$A$4:$I$323')


# ────────────────────────────────────────────
# VERIFICATION
# ────────────────────────────────────────────
print('\n── Verification ──')

remaining = 0
for sn in ['SIGNAL', 'ОТЧЕТ']:
    ws = wb[sn]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and 'ВВОД' in v and v.startswith('='):
                remaining += 1
                cl = get_column_letter(c)
                print(f'  WARNING: {sn}!{cl}{r} still has ВВОД ref')

if remaining == 0:
    print('  OK: No ВВОД references remain in SIGNAL/ОТЧЕТ')

total_names = len(list(wb.defined_names.values()))
print(f'  Total named ranges: {total_names}')


# ────────────────────────────────────────────
# SAVE
# ────────────────────────────────────────────
print(f'\nSaving: {DST}')
wb.save(DST)
print('DONE!')
