import subprocess
subprocess.run(["pip","install","openpyxl"], capture_output=True)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter as cl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

BOLD = Font(bold=True)
BW = Font(bold=True, color="FFFFFF", size=11)
B14W = Font(bold=True, size=14, color="FFFFFF")
B12W = Font(bold=True, size=12, color="FFFFFF")
B11 = Font(bold=True, size=11)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LW = Alignment(horizontal="left", vertical="center", wrap_text=True)
BRD = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))

F_TITLE = PatternFill("solid", fgColor="2F5496")
F_UG = PatternFill("solid", fgColor="D9E2F3")
F_AG = PatternFill("solid", fgColor="E2EFDA")
F_TOT = PatternFill("solid", fgColor="FFF2CC")
F_HDR = PatternFill("solid", fgColor="B4C6E7")
F_WARN = PatternFill("solid", fgColor="FFC7CE")
F_OK = PatternFill("solid", fgColor="C6EFCE")
F_CFG = PatternFill("solid", fgColor="D6DCE4")
F_PAR = PatternFill("solid", fgColor="DAEEF3")
F_FORM = PatternFill("solid", fgColor="F2F2F2")
F_ENTRY = PatternFill("solid", fgColor="FFFFFF")
F_WEEK_A = PatternFill("solid", fgColor="F2F2F2")
F_WEEK_B = PatternFill("solid", fgColor="E8F0FE")
F_PREFILL = PatternFill("solid", fgColor="E8E8E8")
F_PLAN = PatternFill("solid", fgColor="EAF0FB")
F_LVL = PatternFill("solid", fgColor="D9D2E9")

def sc(cell, font=None, fill=None, align=None, border=None, nf=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if nf: cell.number_format = nf

def sr(ws, r1, r2, c1, c2, **kw):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            sc(ws.cell(r,c), **kw)

UNDERGROUND = [
    {"code":"8.1.1","name":"Отделка паркинг и рампы","plan":6600,
     "rw":0.5,"fw":0.5,"rs":[0.65,0.34,0.01],"fs":[0.7,0.2,0.1,0]},
    {"code":"8.1.2","name":"Отделка эвакуац. лестн. клетки подзем.","plan":2530,
     "rw":0.4,"fw":0.6,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.1.3","name":"Лифт. холлы, тамбур-шлюзы подзем.","plan":1800,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.1.4","name":"Технич. помещения подзем.","plan":3200,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.1.5","name":"Прочие помещения подзем.","plan":2000,
     "rw":0.5,"fw":0.5,"rs":[0.2,0.6,0.2],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.1.6","name":"Коммерч. помещения подзем.","plan":2500,
     "rw":0.5,"fw":0.5,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
]
ABOVEGROUND = [
    {"code":"8.2.1","name":"Отделка лобби/гранд-лобби","plan":1500,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],
     "fs":[0.15,0.45,0.15,0.25],"has_furn":True},
    {"code":"8.2.2","name":"Отделка эвакуац. лестн. клетки надзем.","plan":2000,
     "rw":0.4,"fw":0.6,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.3","name":"Лифт. холлы, тамбур-шлюзы надзем.","plan":2500,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.4","name":"Технич. помещения надзем.","plan":4430,
     "rw":0.4,"fw":0.6,"rs":[0.55,0.44,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.5","name":"Прочие помещения надзем.","plan":4000,
     "rw":0.5,"fw":0.5,"rs":[0.2,0.6,0.2],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.6","name":"Коммерч. помещения надзем.","plan":5000,
     "rw":0.5,"fw":0.5,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
    {"code":"8.2.7","name":"Паркинг и рампы надзем.","plan":1500,
     "rw":0.5,"fw":0.5,"rs":[0.3,0.69,0.01],"fs":[0.2,0.6,0.2,0]},
]
ALL = UNDERGROUND + ABOVEGROUND
SURFS = ["Пол","Стены","Потолок"]

dates = []
d = datetime(2026,6,1)
while d <= datetime(2027,4,30):
    dates.append(d)
    d += timedelta(days=7)

DR = 5
LDR = DR + len(dates) - 1
RS = 3
RE = RS + len(ALL) - 1
REF = "СПРАВОЧНИК"
FN = "'ФАКТ - ЧИСТОВАЯ ОТДЕЛКА'"
RN = "'ФАКТ - ЧЕРНОВАЯ ОТДЕЛКА'"
PFN = "'ПЛАН - ЧИСТОВАЯ ОТДЕЛКА'"
PRN = "'ПЛАН - ЧЕРНОВАЯ ОТДЕЛКА'"
VN = "'ВВОД'"

V_DR = 4
BLOCK = 26
V_LAST = V_DR + len(dates) * BLOCK - 1

VCOLS = {"Пол":"E","Стены":"F","Потолок":"G","Мебл.":"H"}

UG_END = RS + len(UNDERGROUND) - 1
AG_START = RS + len(UNDERGROUND)
SUM_HDR = RE + 2
SUM_UG = RE + 3
SUM_AG = RE + 4

def vr(col):
    return f"{col}${V_DR}:{col}${V_LAST}"

def vsum(surf, code, type_name, date_cell=None):
    sc_ = VCOLS[surf]
    parts = f"{VN}!{vr(sc_)}"
    criteria = ""
    if date_cell:
        criteria += f",{VN}!{vr('$A')},{date_cell}"
    criteria += f",{VN}!{vr('$C')},\"{code}\""
    criteria += f",{VN}!{vr('$B')},\"{type_name}\""
    return f"SUMIFS({parts}{criteria})"

def rref(cat):
    return RS + ALL.index(cat)

CFG_COMMENTS = {
    "Статус":"ИСТИНА = карточка активна.\nЛОЖЬ = скрыта.",
    "Гистограмма":"ИСТИНА = гистограмма.\nЛОЖЬ = линейный график.",
    "По месяцам":"ИСТИНА = по месяцам.\nЛОЖЬ = понедельно.",
    "Учитывать дату карточки":"ИСТИНА = фильтр по дате.\nЛОЖЬ = все данные.",
}


# ═══════════════════════════════════════════════════════════
# 1. СПРАВОЧНИК
# ═══════════════════════════════════════════════════════════
ws = wb.active; ws.title = "СПРАВОЧНИК"
ws.sheet_properties.tabColor = "2F5496"

ws.merge_cells("A1:W1")
c = ws["A1"]; c.value = "СПРАВОЧНИК КАТЕГОРИЙ ОТДЕЛКИ"
sc(c,B14W,F_TITLE,CTR); sr(ws,1,1,1,23,fill=F_TITLE)

hdrs = ["Код","Часть","Название","План м²","Вес черн.","Вес чист.","∑ весов",
        "Черн:Пол","Черн:Стены","Черн:Потолок","∑ черн.",
        "Чист:Пол","Чист:Стены","Чист:Потолок","Чист:Мебл.","∑ чист.",
        "Начало","Окончание","Длит.(дни)","Труд-proxy","Вес ур.3 (авто)","Вес ур.3",
        "Глоб. вес"]
for i,h in enumerate(hdrs,1):
    c = ws.cell(2,i,h); sc(c,BOLD,F_HDR,CTR,BRD)

for idx,cat in enumerate(ALL):
    r = RS+idx
    part = "Подзем." if cat["code"].startswith("8.1") else "Надзем."
    fl = F_UG if part=="Подзем." else F_AG
    vals = [cat["code"],part,cat["name"],cat["plan"],
            cat["rw"],cat["fw"],f"=E{r}+F{r}",
            cat["rs"][0],cat["rs"][1],cat["rs"][2],f"=SUM(H{r}:J{r})",
            cat["fs"][0],cat["fs"][1],cat["fs"][2],cat["fs"][3],f"=SUM(L{r}:O{r})"]
    for ci,v in enumerate(vals,1):
        c = ws.cell(r,ci,v); sc(c,border=BRD,fill=fl)
    for cc in [7,11,16]: ws.cell(r,cc).number_format='0.00'

    # Q: Начало (ручной ввод даты)
    sc(ws.cell(r,17),border=BRD,fill=F_ENTRY,nf="DD.MM.YYYY")
    # R: Окончание (ручной ввод даты)
    sc(ws.cell(r,18),border=BRD,fill=F_ENTRY,nf="DD.MM.YYYY")
    # S: Длительность = Окончание - Начало
    ws.cell(r,19,f'=IF(OR(Q{r}="",R{r}=""),0,R{r}-Q{r})')
    sc(ws.cell(r,19),border=BRD,fill=F_FORM,nf='#,##0')
    # T: Труд-proxy = Длительность × План м²
    ws.cell(r,20,f"=S{r}*D{r}")
    sc(ws.cell(r,20),border=BRD,fill=F_FORM,nf='#,##0')
    # U: Вес ур.3 (авто) = Труд / Σ Труд по группе
    ws.cell(r,21,f'=IFERROR(T{r}/SUMIF($B${RS}:$B${RE},$B{r},$T${RS}:$T${RE}),0)')
    sc(ws.cell(r,21),border=BRD,fill=F_FORM,nf='0.0000')
    # V: Вес ур.3 (итоговый, можно перезаписать вручную)
    ws.cell(r,22,f"=U{r}")
    sc(ws.cell(r,22),border=BRD,fill=F_ENTRY,nf='0.0000')
    # W: Глобальный вес = Вес ур.3 × Вес ур.2 группы
    ws.cell(r,23,f'=V{r}*IF($B{r}="Подзем.",$V${SUM_UG},$V${SUM_AG})')
    sc(ws.cell(r,23),border=BRD,fill=F_FORM,nf='0.00%')

ws.cell(2,17).comment=Comment("Введите дату начала работ\nпо категории","Шаблон")
ws.cell(2,18).comment=Comment("Введите дату окончания работ\nпо категории","Шаблон")
ws.cell(2,22).comment=Comment("По умолчанию = авто-расчёт.\nМожно перезаписать вручную\nдля экспертной корректировки.","Шаблон")

for cr_ in [f"G{RS}:G{RE}",f"K{RS}:K{RE}",f"P{RS}:P{RE}"]:
    ws.conditional_formatting.add(cr_,CellIsRule(operator="notEqual",formula=["1"],fill=F_WARN))
    ws.conditional_formatting.add(cr_,CellIsRule(operator="equal",formula=["1"],fill=F_OK))

dv = DataValidation(type="decimal",operator="between",formula1=0,formula2=1,errorTitle="Ошибка",error="0-1")
ws.add_data_validation(dv); dv.add(f"E{RS}:F{RE}"); dv.add(f"H{RS}:J{RE}"); dv.add(f"L{RS}:O{RE}")
dv2 = DataValidation(type="decimal",operator="greaterThanOrEqual",formula1=0,errorTitle="Ошибка",error="≥0")
ws.add_data_validation(dv2); dv2.add(f"D{RS}:D{RE}")

dv_w3 = DataValidation(type="decimal",operator="between",formula1=0,formula2=1,errorTitle="Ошибка",error="Вес 0-1")
ws.add_data_validation(dv_w3); dv_w3.add(f"V{RS}:V{RE}")

ws.column_dimensions["A"].width=8; ws.column_dimensions["B"].width=12; ws.column_dimensions["C"].width=48
ws.column_dimensions["D"].width=14
for x in "EFGHIJKLMNOP": ws.column_dimensions[x].width=13
ws.column_dimensions["Q"].width=14; ws.column_dimensions["R"].width=14
ws.column_dimensions["S"].width=12; ws.column_dimensions["T"].width=14
ws.column_dimensions["U"].width=16; ws.column_dimensions["V"].width=13
ws.column_dimensions["W"].width=12
ws.freeze_panes="A3"
ws.cell(2,15).comment=Comment("Только для 8.2.1 лобби","Шаблон")

# ── Summary rows: Level 2 aggregation ──
ws.merge_cells(f"A{SUM_HDR}:W{SUM_HDR}")
c = ws.cell(SUM_HDR, 1, "АГРЕГАЦИЯ ПО УРОВНЯМ (ур.3 → ур.2 → ур.1)")
sc(c, B12W, F_TITLE, CTR); sr(ws, SUM_HDR, SUM_HDR, 1, 23, fill=F_TITLE)

sum_hdrs = {3:"Название", 4:"План м²", 17:"Начало (мин)", 18:"Окончание (макс)",
            19:"Длит.(дни)", 20:"Труд-proxy", 21:"Вес ур.2 (авто)", 22:"Вес ур.2", 23:"Глоб. вес"}
for ci,h in sum_hdrs.items():
    c = ws.cell(SUM_HDR+1, ci, h); sc(c, BOLD, F_HDR, CTR, BRD)

for sr_row, part_code, part_label, r_start, r_end in [
    (SUM_UG, "8.1", "Подземная часть", RS, UG_END),
    (SUM_AG, "8.2", "Надземная часть", AG_START, RE),
]:
    ws.cell(sr_row, 1, part_code); sc(ws.cell(sr_row, 1), B11, border=BRD)
    ws.cell(sr_row, 3, part_label); sc(ws.cell(sr_row, 3), B11, border=BRD, align=LW)
    ws.cell(sr_row, 4, f"=SUM(D{r_start}:D{r_end})"); sc(ws.cell(sr_row, 4), BOLD, border=BRD, nf='#,##0')
    ws.cell(sr_row, 17, f"=MIN(Q{r_start}:Q{r_end})"); sc(ws.cell(sr_row, 17), border=BRD, nf="DD.MM.YYYY")
    ws.cell(sr_row, 18, f"=MAX(R{r_start}:R{r_end})"); sc(ws.cell(sr_row, 18), border=BRD, nf="DD.MM.YYYY")
    ws.cell(sr_row, 19, f"=IF(OR(Q{sr_row}=0,R{sr_row}=0),0,R{sr_row}-Q{sr_row})")
    sc(ws.cell(sr_row, 19), border=BRD, nf='#,##0')
    ws.cell(sr_row, 20, f"=SUM(T{r_start}:T{r_end})"); sc(ws.cell(sr_row, 20), BOLD, border=BRD, nf='#,##0')
    ws.cell(sr_row, 21, f"=IFERROR(T{sr_row}/(T{SUM_UG}+T{SUM_AG}),0)")
    sc(ws.cell(sr_row, 21), BOLD, border=BRD, nf='0.0000')
    ws.cell(sr_row, 22, f"=U{sr_row}"); sc(ws.cell(sr_row, 22), BOLD, border=BRD, fill=F_ENTRY, nf='0.0000')
    ws.cell(sr_row, 23, f"=V{sr_row}"); sc(ws.cell(sr_row, 23), BOLD, border=BRD, nf='0.00%')
    sr(ws, sr_row, sr_row, 1, 23, border=BRD)

SUM_CHK = SUM_AG + 2
chk_data = [
    (SUM_CHK,   "✓ ∑ вес ур.3 подзем.", f"=SUM(V{RS}:V{UG_END})"),
    (SUM_CHK+1, "✓ ∑ вес ур.3 надзем.", f"=SUM(V{AG_START}:V{RE})"),
    (SUM_CHK+2, "✓ ∑ вес ур.2",         f"=V{SUM_UG}+V{SUM_AG}"),
    (SUM_CHK+3, "✓ ∑ глоб. вес",       f"=SUM(W{RS}:W{RE})"),
]
for cr, label, formula in chk_data:
    ws.cell(cr, 3, label); sc(ws.cell(cr, 3), Font(bold=True, italic=True, size=9), align=LW)
    col = 23 if "глоб" in label else 22
    ws.cell(cr, col, formula); sc(ws.cell(cr, col), nf='0.00%' if col==23 else '0.0000')
    col_letter = "W" if col==23 else "V"
    ws.conditional_formatting.add(f"{col_letter}{cr}", CellIsRule(operator="notEqual", formula=["1"], fill=F_WARN))
    ws.conditional_formatting.add(f"{col_letter}{cr}", CellIsRule(operator="equal", formula=["1"], fill=F_OK))


# ═══════════════════════════════════════════════════════════
# 2. ВВОД
# ═══════════════════════════════════════════════════════════
ws_v = wb.create_sheet("ВВОД")
ws_v.sheet_properties.tabColor = "00B050"

ws_v.merge_cells("A1:I1")
c = ws_v["A1"]; c.value = "ВВОД ДАННЫХ — заполняйте столбцы E-H (белые ячейки)"
sc(c,B14W,F_TITLE,CTR); sr(ws_v,1,1,1,9,fill=F_TITLE)

ws_v.merge_cells("A2:I2")
c = ws_v["A2"]
c.value = "Вводите м². Серые столбцы A-D не редактируйте. Можно также вводить напрямую в ФАКТ-листы."
sc(c,Font(italic=True),F_PAR,CTR)

v_hdrs = ["Дата","Тип","Код","Категория","Пол","Стены","Потолок","Мебл.","✓"]
for i,h in enumerate(v_hdrs,1):
    c = ws_v.cell(3,i,h); sc(c,BOLD,F_HDR,CTR,BRD)

for wi in range(len(dates)):
    block_start = V_DR + wi * BLOCK
    fill = F_WEEK_A if wi%2==0 else F_WEEK_B

    for ti,type_name in enumerate(["Черновая","Чистовая"]):
        for ci,cat in enumerate(ALL):
            r = block_start + ti*13 + ci

            if wi==0 and ti==0 and ci==0:
                ws_v.cell(r,1,dates[0])
                ws_v.cell(r,1).comment=Comment("Первая дата — ручной ввод.\nОстальные автоматически.","Шаблон")
            elif ti==0 and ci==0:
                prev = V_DR + (wi-1)*BLOCK
                ws_v.cell(r,1,f"=A{prev}+7")
            else:
                ws_v.cell(r,1,f"=A{block_start}")
            sc(ws_v.cell(r,1),border=BRD,nf="DD.MM.YYYY",fill=F_PREFILL)

            ws_v.cell(r,2,type_name); sc(ws_v.cell(r,2),border=BRD,fill=F_PREFILL)
            ws_v.cell(r,3,cat["code"]); sc(ws_v.cell(r,3),border=BRD,fill=F_PREFILL)
            ws_v.cell(r,4,cat["name"]); sc(ws_v.cell(r,4),border=BRD,fill=F_PREFILL,align=LW)

            for ec in range(5,9):
                sc(ws_v.cell(r,ec),border=BRD,fill=F_ENTRY,nf='#,##0.00')

            if type_name=="Чистовая" and cat.get("has_furn"):
                ws_v.cell(r,9,f'=COUNTA(E{r}:H{r})&"/4"')
            else:
                ws_v.cell(r,9,f'=COUNTA(E{r}:G{r})&"/3"')
            sc(ws_v.cell(r,9),border=BRD,fill=F_FORM,align=CTR)

ws_v.conditional_formatting.add(
    f"A{V_DR}:I{V_LAST}",
    FormulaRule(formula=[f"AND($A{V_DR}<=TODAY(),$A{V_DR}+6>=TODAY())"],
               fill=PatternFill("solid",fgColor="FFFF99"))
)

dv3 = DataValidation(type="decimal",operator="greaterThanOrEqual",formula1=0,errorTitle="Ошибка",error="≥ 0")
ws_v.add_data_validation(dv3); dv3.add(f"E{V_DR}:H{V_LAST}")

ws_v.column_dimensions["A"].width=12; ws_v.column_dimensions["B"].width=12
ws_v.column_dimensions["C"].width=8; ws_v.column_dimensions["D"].width=38
for x in "EFGH": ws_v.column_dimensions[x].width=12
ws_v.column_dimensions["I"].width=8
ws_v.freeze_panes = f"E{V_DR}"

print(f"ВВОД: {V_LAST} rows ({len(dates)} weeks x {BLOCK})")


# ═══════════════════════════════════════════════════════════
# 3-6. ПЛАН + ФАКТ sheets
# ═══════════════════════════════════════════════════════════
def build_sheet(title, is_finish, tab_color):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color

    col=2; cmap={}
    ug_s=col
    for cat in UNDERGROUND:
        s=col; cols=[col,col+1,col+2]; cmap[cat["code"]]=(s,col+2,cols); col+=3
    ug_e=col-1
    ag_s=col
    for cat in ABOVEGROUND:
        s=col; nc=4 if (is_finish and cat.get("has_furn")) else 3
        cols=list(range(col,col+nc)); cmap[cat["code"]]=(s,col+nc-1,cols); col+=nc
    ag_e=col-1
    col+=1; tug=[col,col+1,col+2]; col+=3
    col+=1; tag=[col,col+1,col+2]; col+=3; maxc=col-1

    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ag_e)
    c=ws.cell(1,1,title.upper())
    sc(c,B14W,F_TITLE,CTR); sr(ws,1,1,1,ag_e,fill=F_TITLE)

    for tc,lbl in [(tug,"Итого подзем."),(tag,"Итого надзем.")]:
        ws.merge_cells(start_row=1,start_column=tc[0],end_row=1,end_column=tc[2])
        c=ws.cell(1,tc[0],lbl); sc(c,BW,F_TITLE,CTR); sr(ws,1,1,tc[0],tc[2],fill=F_TITLE)

    ws.merge_cells(start_row=2,start_column=1,end_row=3,end_column=1)
    c=ws.cell(2,1,"Дата"); sc(c,BOLD,F_HDR,CTR,BRD)
    ws.merge_cells(start_row=2,start_column=ug_s,end_row=2,end_column=ug_e)
    c=ws.cell(2,ug_s,"Подземная часть"); sc(c,BOLD,F_UG,CTR,BRD); sr(ws,2,2,ug_s,ug_e,fill=F_UG,border=BRD)
    ws.merge_cells(start_row=2,start_column=ag_s,end_row=2,end_column=ag_e)
    c=ws.cell(2,ag_s,"Надземная часть"); sc(c,BOLD,F_AG,CTR,BRD); sr(ws,2,2,ag_s,ag_e,fill=F_AG,border=BRD)

    for clist,fl in [(UNDERGROUND,F_UG),(ABOVEGROUND,F_AG)]:
        for cat in clist:
            s,e,_=cmap[cat["code"]]
            if s!=e: ws.merge_cells(start_row=3,start_column=s,end_row=3,end_column=e)
            c=ws.cell(3,s,f"{cat['code']} {cat['name']}"); sc(c,BOLD,fl,CTR,BRD); sr(ws,3,3,s,e,fill=fl,border=BRD)

    for tc in [tug,tag]:
        for i,surf in enumerate(SURFS): c=ws.cell(3,tc[i],surf); sc(c,BOLD,F_TOT,CTR,BRD)

    for cat in ALL:
        _,_,cols=cmap[cat["code"]]; surfs=list(SURFS)
        if is_finish and cat.get("has_furn"): surfs.append("Мебл.")
        fl=F_UG if cat["code"].startswith("8.1") else F_AG
        for i,surf in enumerate(surfs): c=ws.cell(4,cols[i],surf); sc(c,BOLD,fl,CTR,BRD)
    for tc in [tug,tag]: sr(ws,4,4,tc[0],tc[2],fill=F_TOT,border=BRD)

    is_plan = "ПЛАН" in title
    cell_fill = F_PLAN if is_plan else None

    for di in range(len(dates)):
        r=DR+di
        if di==0:
            ws.cell(r,1,dates[0])
            ws.cell(r,1).comment=Comment("Первая дата — ручной ввод.\nОстальные = пред. + 7","Шаблон")
        else: ws.cell(r,1,f"=A{r-1}+7")
        sc(ws.cell(r,1),border=BRD,nf="DD.MM.YYYY")
        for cat in ALL:
            _,_,cols=cmap[cat["code"]]
            for ci in cols:
                sc(ws.cell(r,ci),border=BRD,nf='#,##0.00',fill=cell_fill)
        for si in range(3):
            parts=[f"{cl(cmap[c['code']][2][si])}{r}" for c in UNDERGROUND]
            ws.cell(r,tug[si],f"={'+'.join(parts)}"); sc(ws.cell(r,tug[si]),border=BRD,fill=F_TOT,nf='#,##0.00')
        for si in range(3):
            parts=[f"{cl(cmap[c['code']][2][si])}{r}" for c in ABOVEGROUND]
            ws.cell(r,tag[si],f"={'+'.join(parts)}"); sc(ws.cell(r,tag[si]),border=BRD,fill=F_TOT,nf='#,##0.00')

    ws.conditional_formatting.add(f"A{DR}:A{LDR}",
        FormulaRule(formula=[f"AND($A{DR}<=TODAY(),$A{DR}+6>=TODAY())"],
                    fill=PatternFill("solid",fgColor="FFFF99")))

    ws.column_dimensions["A"].width=12
    for ci in range(2,maxc+1): ws.column_dimensions[cl(ci)].width=10
    ws.freeze_panes=ws.cell(DR,2)
    dv=DataValidation(type="decimal",operator="greaterThanOrEqual",formula1=0,errorTitle="Ошибка",error="≥ 0")
    ws.add_data_validation(dv); dv.add(f"B{DR}:{cl(ag_e)}{LDR}")
    return ws,cmap

ws_plan_rgh, prc = build_sheet("ПЛАН - ЧЕРНОВАЯ ОТДЕЛКА", False, "4472C4")
ws_plan_fin, pfc = build_sheet("ПЛАН - ЧИСТОВАЯ ОТДЕЛКА", True, "4472C4")
ws_rgh, rc = build_sheet("ФАКТ - ЧЕРНОВАЯ ОТДЕЛКА", False, "BF8F00")
ws_fin, fc = build_sheet("ФАКТ - ЧИСТОВАЯ ОТДЕЛКА", True, "548235")


# ═══════════════════════════════════════════════════════════
# 7. ОТЧЕТ
# ═══════════════════════════════════════════════════════════
ws_rep = wb.create_sheet("ОТЧЕТ")
ws_rep.sheet_properties.tabColor = "7030A0"

ws_rep.merge_cells("A1:I1")
c=ws_rep["A1"]; c.value="ОТЧЁТ ПО ОТДЕЛОЧНЫМ РАБОТАМ"
sc(c,B14W,F_TITLE,CTR); sr(ws_rep,1,1,1,9,fill=F_TITLE)

ws_rep.merge_cells("A2:I2")
c=ws_rep["A2"]
c.value="⚠ Автоматический. Собирает из ФАКТ-листов + лист ВВОД. Не вносите данные в оба источника для одной недели."
sc(c,Font(bold=True,italic=True,color="FF0000"),F_WARN,CTR)

rr=4

def write_section(ws, start_r, title, cats):
    r=start_r
    main_rows = {}
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9)
    c=ws.cell(r,1,title); sc(c,B12W,F_TITLE,CTR); sr(ws,r,r,1,9,fill=F_TITLE); r+=1
    ch=["Код","Название","Тип","Вес ур.4","Вес ур.5","План м²","Факт м²","Факт %","Взвеш. %"]
    for i,h in enumerate(ch,1): c=ws.cell(r,i,h); sc(c,BOLD,F_HDR,CTR,BRD)
    r+=1

    for cat in cats:
        rr_=rref(cat); code=cat["code"]
        mr=r
        main_rows[code] = mr
        ws.cell(r,1,code); sc(ws.cell(r,1),B11,border=BRD)
        ws.cell(r,2,f"={REF}!C{rr_}"); sc(ws.cell(r,2),B11,border=BRD,align=LW)
        ws.cell(r,3,"Главная"); sc(ws.cell(r,3),border=BRD)
        sc(ws.cell(r,4),border=BRD); sc(ws.cell(r,5),border=BRD)
        ws.cell(r,6,f"={REF}!D{rr_}"); sc(ws.cell(r,6),border=BRD,fill=F_FORM,nf='#,##0')
        sc(ws.cell(r,7),border=BRD,fill=F_FORM,nf='#,##0.00')
        ws.cell(r,8,f"=IF(F{r}=0,0,G{r}/F{r})"); sc(ws.cell(r,8),border=BRD,fill=F_FORM,nf='0.00%')
        sc(ws.cell(r,9),border=BRD,fill=F_FORM,nf='0.00%')
        r+=1

        phase_rows=[]
        for phase,pname,fsn,wc,col_map in [
            ("r","Черновая (для SIGNAL)",RN,"E",rc),
            ("f","Чистовая (для SIGNAL)",FN,"F",fc),
        ]:
            type_name="Черновая" if phase=="r" else "Чистовая"
            pr=r; phase_rows.append(pr)
            ws.cell(r,2,pname); sc(ws.cell(r,2),BOLD,border=BRD,align=LW)
            sc(ws.cell(r,3),border=BRD)
            ws.cell(r,4,f"={REF}!{wc}{rr_}"); sc(ws.cell(r,4),border=BRD,fill=F_FORM,nf='0.00')
            sc(ws.cell(r,5),border=BRD)
            ws.cell(r,6,f"=F{mr}"); sc(ws.cell(r,6),border=BRD,fill=F_FORM,nf='#,##0')
            sc(ws.cell(r,7),border=BRD,fill=F_FORM,nf='#,##0.00')
            ws.cell(r,8,f"=IF(F{r}=0,0,G{r}/F{r})"); sc(ws.cell(r,8),border=BRD,fill=F_FORM,nf='0.00%')
            ws.cell(r,9,f"=H{r}*D{r}"); sc(ws.cell(r,9),border=BRD,fill=F_FORM,nf='0.00%')
            sc(ws.cell(r,1),border=BRD); r+=1

            _,_,scols=col_map[code]; srows=[]; n_s=len(scols)
            sn=list(SURFS)
            if n_s==4: sn.append("Мебл.")

            for si in range(n_s):
                ws.cell(r,2,f"   {sn[si]}"); sc(ws.cell(r,2),border=BRD,align=LW)
                ws.cell(r,3,"Вид"); sc(ws.cell(r,3),border=BRD)
                sc(ws.cell(r,4),border=BRD)
                sw=cl(8+si) if phase=="r" else cl(12+si)
                ws.cell(r,5,f"={REF}!{sw}{rr_}"); sc(ws.cell(r,5),border=BRD,fill=F_FORM,nf='0.00')
                ws.cell(r,6,f"=F{mr}"); sc(ws.cell(r,6),border=BRD,fill=F_FORM,nf='#,##0')

                fc_=cl(scols[si])
                fact_part=f"SUM({fsn}!{fc_}{DR}:{fc_}{LDR})"
                vvod_part=vsum(sn[si],code,type_name)
                ws.cell(r,7,f"={fact_part}+{vvod_part}")
                sc(ws.cell(r,7),border=BRD,fill=F_FORM,nf='#,##0.00')

                ws.cell(r,8,f"=IF(F{r}=0,0,G{r}/F{r})"); sc(ws.cell(r,8),border=BRD,fill=F_FORM,nf='0.00%')
                ws.cell(r,9,f"=H{r}*E{r}"); sc(ws.cell(r,9),border=BRD,fill=F_FORM,nf='0.00%')
                sc(ws.cell(r,1),border=BRD); srows.append(r); r+=1

            wp=[f"G{s}*E{s}" for s in srows]
            ws.cell(pr,7,f"={'+'.join(wp)}")

        ws.cell(mr,7,f"={'+'.join([f'G{p}*D{p}' for p in phase_rows])}")
        ws.cell(mr,9,f"={'+'.join([f'I{p}' for p in phase_rows])}")

        ws.cell(r,2,"✓ Сумма весов ур.4"); sc(ws.cell(r,2),Font(bold=True,italic=True,size=9),align=LW)
        ws.cell(r,4,f"={'+'.join([f'D{p}' for p in phase_rows])}"); sc(ws.cell(r,4),nf='0.00')
        ws.conditional_formatting.add(f"D{r}",CellIsRule(operator="notEqual",formula=["1"],fill=F_WARN))
        ws.conditional_formatting.add(f"D{r}",CellIsRule(operator="equal",formula=["1"],fill=F_OK))
        r+=2
    return r, main_rows

rr, ag_main = write_section(ws_rep, rr, "НАДЗЕМНАЯ ЧАСТЬ", ABOVEGROUND)
rr, ug_main = write_section(ws_rep, rr, "ПОДЗЕМНАЯ ЧАСТЬ", UNDERGROUND)

ws_rep.column_dimensions["A"].width=8; ws_rep.column_dimensions["B"].width=52; ws_rep.column_dimensions["C"].width=12
for x in "DEFGHI": ws_rep.column_dimensions[x].width=14
ws_rep.freeze_panes="A4"

# ── Level 2 & Level 1 aggregation in ОТЧЕТ ──
rr += 1
ws_rep.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=9)
c = ws_rep.cell(rr, 1, "АГРЕГАЦИЯ: УРОВЕНЬ 3 → УРОВЕНЬ 2 → УРОВЕНЬ 1")
sc(c, B12W, F_TITLE, CTR); sr(ws_rep, rr, rr, 1, 9, fill=F_TITLE); rr += 1

agg_hdrs = ["","Название","","Вес ур.","","План м²","","Факт %","Взвеш. %"]
for i, h in enumerate(agg_hdrs, 1):
    if h: c = ws_rep.cell(rr, i, h); sc(c, BOLD, F_HDR, CTR, BRD)
    else: sc(ws_rep.cell(rr, i), fill=F_HDR, border=BRD)
rr += 1

lvl2_rows = {}
for cats, part_label, part_code, ref_sum_row in [
    (ABOVEGROUND, "8.2 Надземная часть", "8.2", SUM_AG),
    (UNDERGROUND, "8.1 Подземная часть", "8.1", SUM_UG),
]:
    cat_main = ag_main if "8.2" in part_code else ug_main
    lr = rr; lvl2_rows[part_code] = lr

    ws_rep.cell(rr, 1, part_code); sc(ws_rep.cell(rr, 1), font=B11, fill=F_LVL, border=BRD)
    ws_rep.cell(rr, 2, part_label); sc(ws_rep.cell(rr, 2), font=B11, fill=F_LVL, align=LW, border=BRD)
    sc(ws_rep.cell(rr, 3), fill=F_LVL, border=BRD)
    ws_rep.cell(rr, 4, f"={REF}!V{ref_sum_row}"); sc(ws_rep.cell(rr, 4), font=BOLD, fill=F_LVL, border=BRD, nf='0.0000')
    sc(ws_rep.cell(rr, 5), fill=F_LVL, border=BRD)

    plan_parts = [f"F{cat_main[c['code']]}" for c in cats]
    ws_rep.cell(rr, 6, f"={'+'.join(plan_parts)}"); sc(ws_rep.cell(rr, 6), fill=F_LVL, border=BRD, nf='#,##0')
    sc(ws_rep.cell(rr, 7), fill=F_LVL, border=BRD)

    fact_parts = [f"{REF}!V{rref(c)}*H{cat_main[c['code']]}" for c in cats]
    ws_rep.cell(rr, 8, f"={'+'.join(fact_parts)}")
    sc(ws_rep.cell(rr, 8), font=BOLD, fill=F_LVL, border=BRD, nf='0.00%')

    ws_rep.cell(rr, 9, f"=H{rr}*D{rr}"); sc(ws_rep.cell(rr, 9), font=BOLD, fill=F_LVL, border=BRD, nf='0.00%')
    rr += 1

# Weight check for Level 2
ws_rep.cell(rr, 2, "✓ ∑ вес ур.2"); sc(ws_rep.cell(rr, 2), Font(bold=True, italic=True, size=9), align=LW)
l2r = list(lvl2_rows.values())
ws_rep.cell(rr, 4, f"={'+'.join([f'D{r}' for r in l2r])}"); sc(ws_rep.cell(rr, 4), nf='0.0000')
ws_rep.conditional_formatting.add(f"D{rr}", CellIsRule(operator="notEqual", formula=["1"], fill=F_WARN))
ws_rep.conditional_formatting.add(f"D{rr}", CellIsRule(operator="equal", formula=["1"], fill=F_OK))
rr += 2

# Level 1: Overall
F_LVL1 = PatternFill("solid", fgColor="B4A7D6")
ws_rep.cell(rr, 1, "8"); sc(ws_rep.cell(rr, 1), font=B11, fill=F_LVL1, border=BRD)
ws_rep.cell(rr, 2, "ОТДЕЛКА — ОБЩИЙ %"); sc(ws_rep.cell(rr, 2), font=Font(bold=True, size=12), fill=F_LVL1, align=LW, border=BRD)
sc(ws_rep.cell(rr, 3), fill=F_LVL1, border=BRD)
sc(ws_rep.cell(rr, 4), fill=F_LVL1, border=BRD)
sc(ws_rep.cell(rr, 5), fill=F_LVL1, border=BRD)

plan_all = "+".join([f"F{r}" for r in l2r])
ws_rep.cell(rr, 6, f"={plan_all}"); sc(ws_rep.cell(rr, 6), font=Font(bold=True, size=12), fill=F_LVL1, border=BRD, nf='#,##0')
sc(ws_rep.cell(rr, 7), fill=F_LVL1, border=BRD)

fact_all = "+".join([f"I{r}" for r in l2r])
ws_rep.cell(rr, 8, f"={fact_all}")
sc(ws_rep.cell(rr, 8), font=Font(bold=True, size=14), fill=F_LVL1, border=BRD, nf='0.00%')
sc(ws_rep.cell(rr, 9), fill=F_LVL1, border=BRD)


# ═══════════════════════════════════════════════════════════
# 8. SIGNAL — reads ПЛАН + ФАКТ + ВВОД (weighted)
# ═══════════════════════════════════════════════════════════
ws_sig = wb.create_sheet("SIGNAL")
ws_sig.sheet_properties.tabColor = "4472C4"

ws_sig.cell(1,1,"Дата отчета"); sc(ws_sig.cell(1,1),BOLD,F_PAR,LW,BRD)
ws_sig.cell(1,2,datetime(2026,6,29)); sc(ws_sig.cell(1,2),BOLD,F_PAR,border=BRD,nf="DD.MM.YYYY")
ws_sig.cell(1,2).comment=Comment("Замените на =TODAY() / =СЕГОДНЯ()","Шаблон")

def calc_otchet_rows():
    res={}; r=4
    for cats in [ABOVEGROUND,UNDERGROUND]:
        r+=2
        for cat in cats:
            mr=r; r+=1; rr_=r; r+=1; r+=3; fr_=r; r+=1
            n=4 if cat.get("has_furn") else 3; r+=n; r+=2
            res[cat["code"]]={"main":mr,"rough":rr_,"finish":fr_}
    return res
orows=calc_otchet_rows()

sig_r=3
for cat in ALL:
    code=cat["code"]; rr_=rref(cat)

    ws_sig.merge_cells(start_row=sig_r,start_column=1,end_row=sig_r,end_column=19)
    c=ws_sig.cell(sig_r,1,f"═══ {code} {cat['name']} ═══")
    sc(c,B12W,F_TITLE,CTR); sr(ws_sig,sig_r,sig_r,1,19,fill=F_TITLE); sig_r+=1

    cards=[{"name":cat["name"],"type":"total"},
           {"name":f"{cat['name']}: Черновая","type":"rough"},
           {"name":f"{cat['name']}: Чистовая","type":"finish"}]
    cs=[2,8,14]

    cfg=[("Заголовок","План-факт по объемам"),("Статус",True),("Url изображения",""),
         ("Тип","planFact2"),("",""),("Дата","=$B$1"),("Гистограмма",False),
         ("По месяцам",False),("Всего",None),("Тип",None),("Ед. изм.","м2"),
         ("Учитывать дату карточки",False)]

    for ci,cd in enumerate(cards):
        bc=cs[ci]; cr=sig_r
        for fi,(fn,fv) in enumerate(cfg):
            ws_sig.cell(cr,bc,fn); sc(ws_sig.cell(cr,bc),BOLD,F_CFG,LW,BRD)
            if fn=="Всего": ws_sig.cell(cr,bc+1,f"={REF}!D{rr_}")
            elif fn=="Тип" and fi==9: ws_sig.cell(cr,bc+1,cd["name"])
            elif fn=="Дата": ws_sig.cell(cr,bc+1,"=$B$1"); sc(ws_sig.cell(cr,bc+1),nf="DD.MM.YYYY")
            else: ws_sig.cell(cr,bc+1,fv)
            sc(ws_sig.cell(cr,bc+1),border=BRD)
            if fn in CFG_COMMENTS and ci==0:
                ws_sig.cell(cr,bc,fn).comment=Comment(CFG_COMMENTS[fn],"SIGNAL")
            cr+=1
    sig_r+=len(cfg)

    for ci in range(3):
        bc=cs[ci]
        ws_sig.cell(sig_r,bc,"Дата"); sc(ws_sig.cell(sig_r,bc),BOLD,F_HDR,CTR,BRD)
        ws_sig.cell(sig_r,bc+1,"План"); sc(ws_sig.cell(sig_r,bc+1),BOLD,F_HDR,CTR,BRD)
        ws_sig.cell(sig_r,bc+2,"Факт"); sc(ws_sig.cell(sig_r,bc+2),BOLD,F_HDR,CTR,BRD)
    sig_r+=1

    _,_,r_scols=rc[code]; _,_,f_scols=fc[code]
    _,_,pr_scols=prc[code]; _,_,pf_scols=pfc[code]

    for di in range(len(dates)):
        fr=DR+di
        for ci,cd in enumerate(cards):
            bc=cs[ci]; date_cell=f"{cl(bc)}{sig_r}"

            if di==0:
                ws_sig.cell(sig_r,bc,dates[0])
                if ci==0: ws_sig.cell(sig_r,bc).comment=Comment("Первая дата — ручной.\nОстальные = пред.+7","Шаблон")
            else: ws_sig.cell(sig_r,bc,f"={cl(bc)}{sig_r-1}+7")
            sc(ws_sig.cell(sig_r,bc),border=BRD,nf="DD.MM.YYYY")

            # ── ПЛАН (из листов ПЛАН, взвешенный по развесовкам) ──
            if cd["type"]=="rough":
                pp=[]
                for si in range(3):
                    p_=f"{PRN}!{cl(pr_scols[si])}{fr}"
                    w_=f"{REF}!${cl(8+si)}${rr_}"
                    pp.append(f"{p_}*{w_}")
                ws_sig.cell(sig_r,bc+1,f"={'+'.join(pp)}")
            elif cd["type"]=="finish":
                pp=[]
                for si in range(len(pf_scols)):
                    p_=f"{PFN}!{cl(pf_scols[si])}{fr}"
                    w_=f"{REF}!${cl(12+si)}${rr_}"
                    pp.append(f"{p_}*{w_}")
                ws_sig.cell(sig_r,bc+1,f"={'+'.join(pp)}")
            else:
                ws_sig.cell(sig_r,bc+1,f"={cl(cs[1]+1)}{sig_r}+{cl(cs[2]+1)}{sig_r}")
            sc(ws_sig.cell(sig_r,bc+1),border=BRD,fill=F_PLAN,nf='#,##0.00')

            # ── ФАКТ (из листов ФАКТ + ВВОД, взвешенный по развесовкам) ──
            if cd["type"]=="rough":
                parts=[]
                for si in range(3):
                    f_=f"{RN}!{cl(r_scols[si])}{fr}"
                    v_=vsum(SURFS[si],code,"Черновая",date_cell)
                    w_=f"{REF}!${cl(8+si)}${rr_}"
                    parts.append(f"({f_}+{v_})*{w_}")
                ws_sig.cell(sig_r,bc+2,f"={'+'.join(parts)}")
            elif cd["type"]=="finish":
                parts=[]; sn=list(SURFS)
                if len(f_scols)==4: sn.append("Мебл.")
                for si in range(len(f_scols)):
                    f_=f"{FN}!{cl(f_scols[si])}{fr}"
                    v_=vsum(sn[si],code,"Чистовая",date_cell)
                    w_=f"{REF}!${cl(12+si)}${rr_}"
                    parts.append(f"({f_}+{v_})*{w_}")
                ws_sig.cell(sig_r,bc+2,f"={'+'.join(parts)}")
            else:
                ws_sig.cell(sig_r,bc+2,f"={cl(cs[1]+2)}{sig_r}+{cl(cs[2]+2)}{sig_r}")
            sc(ws_sig.cell(sig_r,bc+2),border=BRD,fill=F_FORM,nf='#,##0.00')
        sig_r+=1
    sig_r+=1

for ci in range(1,20): ws_sig.column_dimensions[cl(ci)].width=14
ws_sig.freeze_panes="A3"

# ── Reorder ──
order=["СПРАВОЧНИК","ВВОД",
       "ПЛАН - ЧЕРНОВАЯ ОТДЕЛКА","ПЛАН - ЧИСТОВАЯ ОТДЕЛКА",
       "ФАКТ - ЧЕРНОВАЯ ОТДЕЛКА","ФАКТ - ЧИСТОВАЯ ОТДЕЛКА",
       "ОТЧЕТ","SIGNAL"]
for i,name in enumerate(order):
    idx=wb.sheetnames.index(name); wb.move_sheet(name,offset=i-idx)

# ── Save ──
out=r"\\mr.ru\Service\Personal\ignatov_i\Documents\CloudCode\Отделка SIGNAL\02_Финальная версия - Отделка для SIGNAL v6.xlsx"
wb.save(out)
wb.close()
print(f"OK: {out}")
wb2 = openpyxl.load_workbook(out)
for n in wb2.sheetnames:
    ws=wb2[n]; print(f"  {n}: {ws.max_row}r x {ws.max_column}c")
wb2.close()
