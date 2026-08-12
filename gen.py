"""
Parametrized Excel generator for «Отделка для SIGNAL».
Usage: from gen import generate
       bio = generate(korpus=['К1','К2'], areas={...}, start_date=dt, report_date=dt)
"""

import io
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter as col

PODZEM = [
    ('8.1.1', 'Отделка паркинг и рампы', 6600, .5, .5, [.65, .34, .01], [.70, .20, .10, 0]),
    ('8.1.2', 'Отделка эвакуац. лестн. клетки подзем.', 2530, .4, .6, [.30, .69, .01], [.20, .60, .20, 0]),
    ('8.1.3', 'Лифт. холлы, тамбур-шлюзы подзем.', 1800, .4, .6, [.55, .44, .01], [.20, .60, .20, 0]),
    ('8.1.4', 'Технич. помещения подзем.', 3200, .4, .6, [.55, .44, .01], [.20, .60, .20, 0]),
    ('8.1.5', 'Прочие помещения подзем.', 2000, .5, .5, [.20, .60, .20], [.20, .60, .20, 0]),
    ('8.1.6', 'Коммерч. помещения подзем.', 2500, .5, .5, [.30, .69, .01], [.20, .60, .20, 0]),
]
NADZEM = [
    ('8.2.1', 'Отделка лобби/гранд-лобби', 0, .4, .6, [.55, .44, .01], [.15, .45, .15, .25]),
    ('8.2.2', 'Отделка эвакуац. лестн. клетки надзем.', 0, .4, .6, [.30, .69, .01], [.20, .60, .20, 0]),
    ('8.2.3', 'Лифт. холлы, тамбур-шлюзы надзем.', 0, .4, .6, [.55, .44, .01], [.20, .60, .20, 0]),
    ('8.2.4', 'Технич. помещения надзем.', 0, .4, .6, [.55, .44, .01], [.20, .60, .20, 0]),
    ('8.2.5', 'Прочие помещения надзем.', 0, .5, .5, [.20, .60, .20], [.20, .60, .20, 0]),
    ('8.2.6', 'Коммерч. помещения надзем.', 0, .5, .5, [.30, .69, .01], [.20, .60, .20, 0]),
    ('8.2.7', 'Паркинг и рампы надзем.', 0, .5, .5, [.30, .69, .01], [.20, .60, .20, 0]),
]
DEF_AREAS = [1500, 2000, 2500, 4430, 4000, 5000, 1500]

BOLD = Font(bold=True)
HDR_FILL = PatternFill('solid', fgColor='4472C4')
HDR_FONT = Font(bold=True, color='FFFFFF')
WARN_FONT = Font(italic=True, color='FF0000')
PCT_FMT = '0.0%'
DATE_FMT = 'DD.MM.YYYY'


def _layout(NP, NN, NK):
    p0 = 3; pN = p0 + NP - 1
    n0 = pN + 1; nN = n0 + NN - 1
    ah = nN + 2; a81 = ah + 1; a82 = a81 + 1
    c0 = a82 + 2; c3 = c0 + 3
    kh = c3 + 2; kc = kh + 1; k0 = kc + 1; kN = k0 + NK - 1
    kk = kN + 2; dr = kk + 3
    return dict(p0=p0, pN=pN, n0=n0, nN=nN, ah=ah, a81=a81, a82=a82,
                c0=c0, c3=c3, kh=kh, kc=kc, k0=k0, kN=kN, kk=kk, dr=dr)


def _set_row(ws, r, vals, font=None, fill=None):
    for i, v in enumerate(vals, 1):
        if v is None:
            continue
        c = ws.cell(r, i, v)
        if font: c.font = font
        if fill: c.fill = fill


def _fact_cl(ci, si):
    return col(2 + ci * 3 + si)


def _chern_wc(si):
    return col(8 + si)


def _chist_wc(si):
    return col(12 + si)


def _spr_row(ci, is_ndz, S):
    return (S['n0'] if is_ndz else S['p0']) + ci


def _spravochnik(wb, S, korpus, areas, NP, NN, NK, start_date):
    ws = wb.create_sheet('СПРАВОЧНИК')
    ws.cell(1, 1, 'СПРАВОЧНИК КАТЕГОРИЙ ОТДЕЛКИ').font = BOLD
    headers = [
        'Код', 'Часть', 'Название', 'План м²', 'Вес черн.', 'Вес чист.', '∑ весов',
        'Черн:Пол', 'Черн:Стены', 'Черн:Потолок', '∑ черн.',
        'Чист:Пол', 'Чист:Стены', 'Чист:Потолок', 'Чист:Мебл.', '∑ чист.',
        'Начало', 'Окончание', 'Длит.(дни)', 'Труд-proxy',
        'Вес ур.3 (авто)', 'Вес ур.3', 'Глоб. вес',
    ]
    _set_row(ws, 2, headers, font=HDR_FONT, fill=HDR_FILL)

    def write_cat(r, code, part, name, plan, wc, wch, cw, chw):
        ws.cell(r, 1, code); ws.cell(r, 2, part); ws.cell(r, 3, name)
        ws.cell(r, 4, plan); ws.cell(r, 5, wc); ws.cell(r, 6, wch)
        ws.cell(r, 7, f'=E{r}+F{r}')
        for i, w in enumerate(cw): ws.cell(r, 8 + i, w)
        ws.cell(r, 11, f'=SUM(H{r}:J{r})')
        for i, w in enumerate(chw): ws.cell(r, 12 + i, w)
        ws.cell(r, 16, f'=SUM(L{r}:O{r})')
        ws.cell(r, 19, f'=IF(OR(Q{r}="",R{r}=""),0,R{r}-Q{r})')
        ws.cell(r, 20, f'=S{r}*D{r}')
        nN = S['nN']
        ws.cell(r, 21, f'=IFERROR(T{r}/SUMIF($B$3:$B${nN},$B{r},$T$3:$T${nN}),0)')
        ws.cell(r, 22, f'=U{r}')
        a81, a82 = S['a81'], S['a82']
        ws.cell(r, 23, f'=V{r}*IF($B{r}="Подзем.",$V${a81},$V${a82})')

    for i, (code, name, plan, wc, wch, cw, chw) in enumerate(PODZEM):
        write_cat(S['p0'] + i, code, 'Подзем.', name, plan, wc, wch, cw, chw)

    for i, (code, name, _, wc, wch, cw, chw) in enumerate(NADZEM):
        r = S['n0'] + i
        ac = col(2 + i)
        plan_f = '=' + '+'.join(f'{ac}{S["k0"] + ki}' for ki in range(NK))
        write_cat(r, code, 'Надзем.', name, plan_f, wc, wch, cw, chw)

    ws.cell(S['ah'], 1, 'АГРЕГАЦИЯ ПО УРОВНЯМ (ур.3 → ур.2 → ур.1)').font = BOLD
    for tag, ar, rs, re in [('8.1', S['a81'], S['p0'], S['pN']),
                             ('8.2', S['a82'], S['n0'], S['nN'])]:
        label = 'Подземная часть' if tag == '8.1' else 'Надземная часть'
        ws.cell(ar, 1, tag); ws.cell(ar, 3, label)
        ws.cell(ar, 4, f'=SUM(D{rs}:D{re})')
        ws.cell(ar, 17, f'=MIN(Q{rs}:Q{re})')
        ws.cell(ar, 18, f'=MAX(R{rs}:R{re})')
        ws.cell(ar, 19, f'=IF(OR(Q{ar}=0,R{ar}=0),0,R{ar}-Q{ar})')
        ws.cell(ar, 20, f'=SUM(T{rs}:T{re})')
        a81, a82 = S['a81'], S['a82']
        ws.cell(ar, 21, f'=IFERROR(T{ar}/(T{a81}+T{a82}),0)')
        ws.cell(ar, 22, f'=U{ar}'); ws.cell(ar, 23, f'=V{ar}')

    ws.cell(S['c0'], 3, '✓ ∑ вес ур.3 подзем.')
    ws.cell(S['c0'], 22, f'=SUM(V{S["p0"]}:V{S["pN"]})')
    ws.cell(S['c0'] + 1, 3, '✓ ∑ вес ур.3 надзем.')
    ws.cell(S['c0'] + 1, 22, f'=SUM(V{S["n0"]}:V{S["nN"]})')
    ws.cell(S['c0'] + 2, 3, '✓ ∑ вес ур.2')
    ws.cell(S['c0'] + 2, 22, f'=V{S["a81"]}+V{S["a82"]}')
    ws.cell(S['c3'], 3, '✓ ∑ глоб. вес')
    ws.cell(S['c3'], 23, f'=SUM(W{S["p0"]}:W{S["nN"]})')

    ws.cell(S['kh'], 1, 'РАСПРЕДЕЛЕНИЕ НАДЗЕМНОЙ ЧАСТИ ПО КОРПУСАМ').font = BOLD
    khd = ['Корпус'] + [c[0] for c in NADZEM] + ['Σ план', 'Вес корп.(авто)', 'Вес корп.']
    _set_row(ws, S['kc'], khd, font=BOLD)

    sc = 2 + NN
    sr = '+'.join(f'{col(sc)}{S["k0"] + ki}' for ki in range(NK))
    for ki, kname in enumerate(korpus):
        r = S['k0'] + ki
        ws.cell(r, 1, kname)
        ka = areas.get(kname, DEF_AREAS)
        for ci, a in enumerate(ka):
            ws.cell(r, 2 + ci, a)
        la = col(2 + NN - 1)
        ws.cell(r, sc, f'=SUM(B{r}:{la}{r})')
        ws.cell(r, sc + 1, f'=IFERROR({col(sc)}{r}/({sr}),0)')
        ws.cell(r, sc + 2, f'={col(sc + 1)}{r}')

    ws.cell(S['kk'], 1, '✓ ∑ вес корп.')
    ws.cell(S['kk'], sc + 2, '=' + '+'.join(f'{col(sc + 2)}{S["k0"] + ki}' for ki in range(NK)))

    ws.cell(S['dr'], 1, 'Дата начала проекта')
    ws.cell(S['dr'], 2, start_date)
    ws.cell(S['dr'], 2).number_format = DATE_FMT
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['A'].width = 8


def _vvod(wb, S, suffix, cats, weeks):
    sn = f'ВВОД {suffix}'
    ws = wb.create_sheet(sn)
    ws.cell(1, 1, f'ВВОД ДАННЫХ — {sn}').font = BOLD
    ws.cell(2, 1, 'Вводите м². Серые столбцы A-D не редактируйте.').font = WARN_FONT
    _set_row(ws, 3, ['Дата', 'Тип', 'Код', 'Категория', 'Пол', 'Стены', 'Потолок', 'Мебл.', '✓'], font=BOLD)
    rpw = len(cats) * 2
    dr = S['dr']
    for w in range(weeks):
        ws0 = 4 + w * rpw
        for ti, tn in enumerate(['Черновая', 'Чистовая']):
            for ci, (code, name, *_) in enumerate(cats):
                r = ws0 + ti * len(cats) + ci
                if w == 0 and ti == 0 and ci == 0:
                    ws.cell(r, 1, f'=СПРАВОЧНИК!$B${dr}')
                elif ti == 0 and ci == 0:
                    ws.cell(r, 1, f'=A{4 + (w - 1) * rpw}+7')
                else:
                    ws.cell(r, 1, f'=A{ws0}')
                ws.cell(r, 1).number_format = DATE_FMT
                ws.cell(r, 2, tn); ws.cell(r, 3, code); ws.cell(r, 4, name)
                ws.cell(r, 9, f'=COUNTA(E{r}:G{r})&"/3"')
    ws.column_dimensions['D'].width = 40


def _plan_fact(wb, S, kind, cc, psuf, cats, weeks, has_mebl=False):
    sn = f'{kind} {cc} {psuf}'
    ws = wb.create_sheet(sn)
    nc = len(cats)
    gc = 2 + nc * 3; ts = gc + 1
    nt = 4 if has_mebl else 3
    tl = ['Пол', 'Стены', 'Потолок'] + (['Мебл.'] if has_mebl else [])

    ws.cell(1, 1, sn).font = BOLD
    ws.cell(1, ts, f'Итого {psuf.lower()}')
    ws.cell(2, 1, 'Дата')
    ws.cell(2, 2, 'Подземная часть' if '8.1' in cats[0][0] else 'Надземная часть').font = BOLD
    for ci, (code, name, *_) in enumerate(cats):
        ws.cell(3, 2 + ci * 3, f'{code} {name}').font = BOLD
    for ci in range(nc):
        for si, sn_ in enumerate(['Пол', 'Стены', 'Потолок']):
            ws.cell(4, 2 + ci * 3 + si, sn_)
    for ti, tl_ in enumerate(tl):
        ws.cell(3, ts + ti, tl_).font = BOLD

    dr = S['dr']
    for w in range(weeks):
        r = 5 + w
        ws.cell(r, 1, f'=СПРАВОЧНИК!$B${dr}' if w == 0 else f'=A{r - 1}+7')
        ws.cell(r, 1).number_format = DATE_FMT
        for si in range(min(3, nt)):
            src = '+'.join(f'{col(2 + ci * 3 + si)}{r}' for ci in range(nc))
            ws.cell(r, ts + si, f'={src}')
    ws.column_dimensions['A'].width = 12


def _otchet(wb, S, korpus, weeks, NP, NN, NK):
    ws = wb.create_sheet('ОТЧЕТ')
    ws.cell(1, 1, 'ОТЧЁТ ПО ОТДЕЛОЧНЫМ РАБОТАМ').font = BOLD
    ws.cell(2, 1, '⚠ Автоматический. Не вносите данные напрямую.').font = WARN_FONT
    hdr = ['Код', 'Название', 'Тип', 'Вес ур.4', 'Вес ур.5', 'План м²', 'Факт м²', 'Факт %', 'Взвеш. %']
    r = 4

    def wsec(r, title, cats, is_ndz, fp, pfn):
        ws.cell(r, 1, title).font = BOLD; r += 1
        _set_row(ws, r, hdr, font=BOLD); r += 1
        for ci, (code, name, plan, wc, wch, cw, chw) in enumerate(cats):
            sr = _spr_row(ci, is_ndz, S); m = r
            ws.cell(r, 1, code); ws.cell(r, 2, f'=СПРАВОЧНИК!C{sr}')
            ws.cell(r, 3, 'Главная'); ws.cell(r, 6, pfn(ci))
            ws.cell(r, 7, f'=G{r+1}*D{r+1}+G{r+5}*D{r+5}')
            ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})'); ws.cell(r, 8).number_format = PCT_FMT
            ws.cell(r, 9, f'=I{r+1}+I{r+5}'); ws.cell(r, 9).number_format = PCT_FMT; r += 1

            ws.cell(r, 2, 'Черновая'); ws.cell(r, 4, f'=СПРАВОЧНИК!E{sr}')
            ws.cell(r, 6, f'=F{m}')
            ws.cell(r, 7, f'=G{r+1}*E{r+1}+G{r+2}*E{r+2}+G{r+3}*E{r+3}')
            ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})'); ws.cell(r, 8).number_format = PCT_FMT
            ws.cell(r, 9, f'=H{r}*D{r}'); ws.cell(r, 9).number_format = PCT_FMT; r += 1

            fsc = f'ФАКТ ЧЕРН {fp}'
            for si, sn_ in enumerate(['Пол', 'Стены', 'Потолок']):
                fc = _fact_cl(ci, si)
                ws.cell(r, 2, f'   {sn_}'); ws.cell(r, 3, 'Вид')
                ws.cell(r, 5, f'=СПРАВОЧНИК!{_chern_wc(si)}{sr}')
                ws.cell(r, 6, f'=F{m}')
                ws.cell(r, 7, f"=SUM('{fsc}'!{fc}5:{fc}{4+weeks})")
                ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})'); ws.cell(r, 8).number_format = PCT_FMT
                ws.cell(r, 9, f'=H{r}*E{r}'); ws.cell(r, 9).number_format = PCT_FMT; r += 1

            cr = r
            ws.cell(r, 2, 'Чистовая'); ws.cell(r, 4, f'=СПРАВОЧНИК!F{sr}')
            ws.cell(r, 6, f'=F{m}')
            ws.cell(r, 7, f'=G{r+1}*E{r+1}+G{r+2}*E{r+2}+G{r+3}*E{r+3}')
            ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})'); ws.cell(r, 8).number_format = PCT_FMT
            ws.cell(r, 9, f'=H{r}*D{r}'); ws.cell(r, 9).number_format = PCT_FMT; r += 1

            fsi = f'ФАКТ ЧИСТ {fp}'
            for si, sn_ in enumerate(['Пол', 'Стены', 'Потолок']):
                fc = _fact_cl(ci, si)
                ws.cell(r, 2, f'   {sn_}'); ws.cell(r, 3, 'Вид')
                ws.cell(r, 5, f'=СПРАВОЧНИК!{_chist_wc(si)}{sr}')
                ws.cell(r, 6, f'=F{m}')
                ws.cell(r, 7, f"=SUM('{fsi}'!{fc}5:{fc}{4+weeks})")
                ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})'); ws.cell(r, 8).number_format = PCT_FMT
                ws.cell(r, 9, f'=H{r}*E{r}'); ws.cell(r, 9).number_format = PCT_FMT; r += 1

            ws.cell(r, 2, '✓ Сумма весов ур.4')
            ws.cell(r, 4, f'=D{m+1}+D{cr}'); r += 2
        return r

    r = wsec(r, 'ПОДЗЕМНАЯ ЧАСТЬ', PODZEM, False, 'ПОДЗЕМ',
             lambda ci: f'=СПРАВОЧНИК!D{_spr_row(ci, False, S)}')
    for ki, kn in enumerate(korpus):
        kr = S['k0'] + ki
        r = wsec(r, f'КОРПУС {kn}', NADZEM, True, kn,
                 lambda ci, _kr=kr: f'=СПРАВОЧНИК!{col(2 + ci)}{_kr}')
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['A'].width = 8


def _signal_range_name(label, korpus):
    for kn in korpus:
        if kn in label:
            if 'Общий' in label: return f'{kn}_Общий_8_2'
            if 'Черновая' in label: return f'{kn}_Черновая_8_2'
            if 'Чистовая' in label: return f'{kn}_Чистовая_8_2'
    if '8.1' in label:
        if 'Общий' in label: return 'Общий_8_1'
        if 'Черновая' in label: return 'Черновая_8_1'
        if 'Чистовая' in label: return 'Чистовая_8_1'
    if '8.2' in label and 'Надземная' in label:
        if 'Общий' in label: return 'Общий_8_2'
        if 'Черновая' in label: return 'Черновая_8_2'
        if 'Чистовая' in label: return 'Чистовая_8_2'
    if 'Отделка общий' in label:
        if 'Общий' in label: return 'Общая'
        if 'Черновая' in label: return 'Черновая'
        if 'Чистовая' in label: return 'Чистовая'
    return label.replace(' ', '_').replace(':', '').replace('.', '_')


def _signal(wb, S, korpus, weeks, NN, NK, report_date):
    ws = wb.create_sheet('SIGNAL')
    blocks = []

    def ab(label, btype, **kw):
        idx = len(blocks)
        b = {'idx': idx, 'base': 2 + idx * 6, 'label': label, 'type': btype}
        b.update(kw)
        blocks.append(b)
        return b

    bpo = ab('8.1 Подземная часть: Общий', 'combined')
    bpc = ab('8.1 Подземная часть: Черновая', 'weighted',
             sheet='ПЛАН ЧЕРН ПОДЗЕМ', fsheet='ФАКТ ЧЕРН ПОДЗЕМ',
             cats=PODZEM, is_nadzem=False, wtype='chern')
    bpi = ab('8.1 Подземная часть: Чистовая', 'weighted',
             sheet='ПЛАН ЧИСТ ПОДЗЕМ', fsheet='ФАКТ ЧИСТ ПОДЗЕМ',
             cats=PODZEM, is_nadzem=False, wtype='chist')
    bpo['chern'] = bpc; bpo['chist'] = bpi
    bpo['vsego'] = f'=СПРАВОЧНИК!D{S["a81"]}'

    ko_l, kc_l, ki_l = [], [], []
    for ki, kn in enumerate(korpus):
        bko = ab(f'8.2 {kn}: Общий', 'combined')
        bkc = ab(f'8.2 {kn}: Черновая', 'weighted',
                 sheet=f'ПЛАН ЧЕРН {kn}', fsheet=f'ФАКТ ЧЕРН {kn}',
                 cats=NADZEM, is_nadzem=True, wtype='chern')
        bki = ab(f'8.2 {kn}: Чистовая', 'weighted',
                 sheet=f'ПЛАН ЧИСТ {kn}', fsheet=f'ФАКТ ЧИСТ {kn}',
                 cats=NADZEM, is_nadzem=True, wtype='chist')
        bko['chern'] = bkc; bko['chist'] = bki
        sc = 2 + NN
        bko['vsego'] = f'=СПРАВОЧНИК!{col(sc)}{S["k0"] + ki}'
        ko_l.append(bko); kc_l.append(bkc); ki_l.append(bki)

    bno = ab('8.2 Надземная часть: Общий', 'sum_blocks', sources=ko_l)
    bnc = ab('8.2 Надземная часть: Черновая', 'sum_blocks', sources=kc_l)
    bni = ab('8.2 Надземная часть: Чистовая', 'sum_blocks', sources=ki_l)
    for b in (bno, bnc, bni):
        b['vsego'] = f'=СПРАВОЧНИК!D{S["a82"]}'

    bto = ab('8. Отделка общий: Общий', 'sum_two', a=bpo, b=bno)
    btc = ab('8. Отделка общий: Черновая', 'sum_two', a=bpc, b=bnc)
    bti = ab('8. Отделка общий: Чистовая', 'sum_two', a=bpi, b=bni)
    for b in (bto, btc, bti):
        b['vsego'] = f'=СПРАВОЧНИК!D{S["a81"]}+СПРАВОЧНИК!D{S["a82"]}'

    ws.cell(1, 1, 'Дата отчета')
    ws.cell(2, 1, report_date)
    ws.cell(2, 1).number_format = DATE_FMT

    dr = S['dr']
    for b in blocks:
        bc = b['base']; dc = bc + 1; pc = bc + 2; fc = bc + 3
        ws.cell(1, bc, 'Заголовок'); ws.cell(1, dc, 'План-факт по объемам')
        ws.cell(2, bc, 'Статус'); ws.cell(2, dc, 'true')
        ws.cell(3, bc, 'Url изображения')
        ws.cell(4, bc, 'Тип'); ws.cell(4, dc, 'planFact2')
        ws.cell(6, bc, 'Дата'); ws.cell(6, dc, '=$A$2')
        ws.cell(7, bc, 'Гистограмма'); ws.cell(7, dc, 'false')
        ws.cell(8, bc, 'По месяцам'); ws.cell(8, dc, 'false')
        ws.cell(9, bc, 'Всего'); ws.cell(9, dc, b.get('vsego', ''))
        ws.cell(10, bc, 'Тип'); ws.cell(10, dc, b['label'])
        ws.cell(11, bc, 'Ед. изм.'); ws.cell(11, dc, 'м2')
        ws.cell(12, bc, 'Учитывать дату карточки'); ws.cell(12, dc, 'true')
        ws.cell(13, bc, 'Данные')

        for w in range(weeks):
            r = 14 + w; pr = 5 + w
            if w == 0:
                ws.cell(r, dc, f'=СПРАВОЧНИК!$B${dr}')
            else:
                ws.cell(r, dc, f'={col(dc)}{r - 1}+7')
            ws.cell(r, dc).number_format = DATE_FMT
            bt = b['type']

            if bt == 'weighted':
                nc = len(b['cats']); nd = b['is_nadzem']
                wcf = _chist_wc if b['wtype'] == 'chist' else _chern_wc
                pt, ft = [], []
                for ci in range(nc):
                    sr = _spr_row(ci, nd, S)
                    for si in range(3):
                        pcl = _fact_cl(ci, si); wc = wcf(si)
                        pt.append(f"'{b['sheet']}'!{pcl}{pr}*СПРАВОЧНИК!${wc}${sr}")
                        ft.append(f"'{b['fsheet']}'!{pcl}{pr}*СПРАВОЧНИК!${wc}${sr}")
                ws.cell(r, pc, '=' + '+'.join(pt))
                ws.cell(r, fc, '=' + '+'.join(ft))
            elif bt == 'combined':
                chp = b['chern']['base'] + 2; chip = b['chist']['base'] + 2
                chf = b['chern']['base'] + 3; chif = b['chist']['base'] + 3
                ws.cell(r, pc, f'={col(chp)}{r}+{col(chip)}{r}')
                ws.cell(r, fc, f'={col(chf)}{r}+{col(chif)}{r}')
            elif bt == 'sum_blocks':
                ws.cell(r, pc, '=' + '+'.join(f'{col(s["base"]+2)}{r}' for s in b['sources']))
                ws.cell(r, fc, '=' + '+'.join(f'{col(s["base"]+3)}{r}' for s in b['sources']))
            elif bt == 'sum_two':
                ws.cell(r, pc, f'={col(b["a"]["base"]+2)}{r}+{col(b["b"]["base"]+2)}{r}')
                ws.cell(r, fc, f'={col(b["a"]["base"]+3)}{r}+{col(b["b"]["base"]+3)}{r}')

    sig_ranges = []
    for b in blocks:
        c1 = col(b['base']); c2 = col(b['base'] + 3)
        name = _signal_range_name(b['label'], korpus)
        sig_ranges.append((name, f"'SIGNAL'!${c1}:${c2}"))
    return sig_ranges


def generate(korpus, areas=None, start_date=None, report_date=None, weeks=48):
    """Generate XLSX workbook. Returns BytesIO ready for send_file()."""
    from datetime import datetime
    if areas is None:
        areas = {k: DEF_AREAS[:] for k in korpus}
    if start_date is None:
        start_date = datetime(2026, 6, 1)
    if report_date is None:
        report_date = datetime(2026, 6, 29)

    NP, NN, NK = len(PODZEM), len(NADZEM), len(korpus)
    S = _layout(NP, NN, NK)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _spravochnik(wb, S, korpus, areas, NP, NN, NK, start_date)

    _vvod(wb, S, 'ПОДЗЕМ', PODZEM, weeks)
    for k in korpus:
        _vvod(wb, S, k, NADZEM, weeks)

    parts = [('ПОДЗЕМ', PODZEM, False)] + [(k, NADZEM, True) for k in korpus]
    for suf, cats, nd in parts:
        for kind in ['ПЛАН', 'ФАКТ']:
            _plan_fact(wb, S, kind, 'ЧЕРН', suf, cats, weeks)
            _plan_fact(wb, S, kind, 'ЧИСТ', suf, cats, weeks, has_mebl=nd)

    _otchet(wb, S, korpus, weeks, NP, NN, NK)
    sig_ranges = _signal(wb, S, korpus, weeks, NN, NK, report_date)
    wb.create_sheet('Лист2')

    for name, ref in sig_ranges:
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
