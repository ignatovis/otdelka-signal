#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор файла «Отделка для SIGNAL».
Настраиваемое количество корпусов — меняйте KORPUS и KORPUS_AREAS.

python generate_otdelka.py
"""

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as col
from datetime import datetime
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ================================================================
#  КОНФИГУРАЦИЯ — редактируйте здесь
# ================================================================

OUTPUT = 'Отделка SIGNAL/Отделка для SIGNAL (generated).xlsx'

START_DATE = datetime(2026, 6, 1)
REPORT_DATE = datetime(2026, 6, 29)
WEEKS = 48

# ← Главный параметр: список корпусов
KORPUS = ['К1', 'К2', 'К3']

# Площади (м²) по категориям 8.2.1–8.2.7 для каждого корпуса
KORPUS_AREAS = {
    'К1': [1500, 2000, 2500, 4430, 4000, 5000, 1500],
    'К2': [1500, 2000, 2500, 4430, 4000, 5000, 1500],
    'К3': [1500, 2000, 2500, 4430, 4000, 5000, 1500],
}

# (код, название, план_м2, вес_черн, вес_чист, [черн_пол,стены,потолок], [чист_пол,стены,потолок,мебл])
PODZEM = [
    ('8.1.1','Отделка паркинг и рампы',           6600,.5,.5,[.65,.34,.01],[.70,.20,.10,0]),
    ('8.1.2','Отделка эвакуац. лестн. клетки подзем.',2530,.4,.6,[.30,.69,.01],[.20,.60,.20,0]),
    ('8.1.3','Лифт. холлы, тамбур-шлюзы подзем.', 1800,.4,.6,[.55,.44,.01],[.20,.60,.20,0]),
    ('8.1.4','Технич. помещения подзем.',           3200,.4,.6,[.55,.44,.01],[.20,.60,.20,0]),
    ('8.1.5','Прочие помещения подзем.',            2000,.5,.5,[.20,.60,.20],[.20,.60,.20,0]),
    ('8.1.6','Коммерч. помещения подзем.',          2500,.5,.5,[.30,.69,.01],[.20,.60,.20,0]),
]

NADZEM = [
    ('8.2.1','Отделка лобби/гранд-лобби',          0,.4,.6,[.55,.44,.01],[.15,.45,.15,.25]),
    ('8.2.2','Отделка эвакуац. лестн. клетки надзем.',0,.4,.6,[.30,.69,.01],[.20,.60,.20,0]),
    ('8.2.3','Лифт. холлы, тамбур-шлюзы надзем.',  0,.4,.6,[.55,.44,.01],[.20,.60,.20,0]),
    ('8.2.4','Технич. помещения надзем.',            0,.4,.6,[.55,.44,.01],[.20,.60,.20,0]),
    ('8.2.5','Прочие помещения надзем.',             0,.5,.5,[.20,.60,.20],[.20,.60,.20,0]),
    ('8.2.6','Коммерч. помещения надзем.',           0,.5,.5,[.30,.69,.01],[.20,.60,.20,0]),
    ('8.2.7','Паркинг и рампы надзем.',              0,.5,.5,[.30,.69,.01],[.20,.60,.20,0]),
]

# ================================================================
#  LAYOUT CONSTANTS (computed)
# ================================================================

NP = len(PODZEM)
NN = len(NADZEM)
NK = len(KORPUS)


class S:
    """СПРАВОЧНИК row layout."""
    title = 1
    hdr = 2
    p0 = 3
    pN = p0 + NP - 1
    n0 = pN + 1
    nN = n0 + NN - 1
    agg_hdr = nN + 2
    agg81 = agg_hdr + 1
    agg82 = agg81 + 1
    chk0 = agg82 + 2
    chk3 = chk0 + 3
    korp_hdr = chk3 + 2
    korp_col = korp_hdr + 1
    k0 = korp_col + 1
    kN = k0 + NK - 1
    kchk = kN + 2
    date_r = kchk + 3


# Styles
BOLD = Font(bold=True)
HDR_FILL = PatternFill('solid', fgColor='4472C4')
HDR_FONT = Font(bold=True, color='FFFFFF')
WARN_FONT = Font(italic=True, color='FF0000')
PCT_FMT = '0.0%'
DATE_FMT = 'DD.MM.YYYY'

# ================================================================
#  HELPERS
# ================================================================


def set_row(ws, r, vals, font=None, fill=None):
    for i, v in enumerate(vals, 1):
        if v is None:
            continue
        c = ws.cell(r, i, v)
        if font:
            c.font = font
        if fill:
            c.fill = fill


def sprav_cat_row(cat_idx, is_nadzem):
    return (S.n0 if is_nadzem else S.p0) + cat_idx


def fact_col_letter(cat_idx, surf_idx):
    return col(2 + cat_idx * 3 + surf_idx)


def chern_weight_col(surf_idx):
    return col(8 + surf_idx)


def chist_weight_col(surf_idx):
    return col(12 + surf_idx)


# ================================================================
#  СПРАВОЧНИК
# ================================================================

def build_spravochnik(wb):
    ws = wb.create_sheet('СПРАВОЧНИК')
    print('  СПРАВОЧНИК')

    ws.cell(S.title, 1, 'СПРАВОЧНИК КАТЕГОРИЙ ОТДЕЛКИ').font = BOLD

    headers = [
        'Код', 'Часть', 'Название', 'План м²', 'Вес черн.', 'Вес чист.', '∑ весов',
        'Черн:Пол', 'Черн:Стены', 'Черн:Потолок', '∑ черн.',
        'Чист:Пол', 'Чист:Стены', 'Чист:Потолок', 'Чист:Мебл.', '∑ чист.',
        'Начало', 'Окончание', 'Длит.(дни)', 'Труд-proxy',
        'Вес ур.3 (авто)', 'Вес ур.3', 'Глоб. вес',
    ]
    set_row(ws, S.hdr, headers, font=HDR_FONT, fill=HDR_FILL)

    def write_cat(r, code, part_label, name, plan, wc, wch, cw, chw):
        ws.cell(r, 1, code)
        ws.cell(r, 2, part_label)
        ws.cell(r, 3, name)
        ws.cell(r, 4, plan)
        ws.cell(r, 5, wc)
        ws.cell(r, 6, wch)
        ws.cell(r, 7, f'=E{r}+F{r}')
        for i, w in enumerate(cw):
            ws.cell(r, 8 + i, w)
        ws.cell(r, 11, f'=SUM(H{r}:J{r})')
        for i, w in enumerate(chw):
            ws.cell(r, 12 + i, w)
        ws.cell(r, 16, f'=SUM(L{r}:O{r})')
        ws.cell(r, 19, f'=IF(OR(Q{r}="",R{r}=""),0,R{r}-Q{r})')
        ws.cell(r, 20, f'=S{r}*D{r}')
        ws.cell(r, 21, f'=IFERROR(T{r}/SUMIF($B$3:$B${S.nN},$B{r},$T$3:$T${S.nN}),0)')
        ws.cell(r, 22, f'=U{r}')
        ws.cell(r, 23, f'=V{r}*IF($B{r}="Подзем.",$V${S.agg81},$V${S.agg82})')

    for i, (code, name, plan, wc, wch, cw, chw) in enumerate(PODZEM):
        write_cat(S.p0 + i, code, 'Подзем.', name, plan, wc, wch, cw, chw)

    for i, (code, name, _, wc, wch, cw, chw) in enumerate(NADZEM):
        r = S.n0 + i
        area_col = col(2 + i)
        plan_formula = '=' + '+'.join(f'{area_col}{S.k0 + ki}' for ki in range(NK))
        write_cat(r, code, 'Надзем.', name, plan_formula, wc, wch, cw, chw)

    # Aggregation
    ws.cell(S.agg_hdr, 1, 'АГРЕГАЦИЯ ПО УРОВНЯМ (ур.3 → ур.2 → ур.1)').font = BOLD
    for tag, ar, rs, re in [('8.1', S.agg81, S.p0, S.pN), ('8.2', S.agg82, S.n0, S.nN)]:
        label = 'Подземная часть' if tag == '8.1' else 'Надземная часть'
        ws.cell(ar, 1, tag)
        ws.cell(ar, 3, label)
        ws.cell(ar, 4, f'=SUM(D{rs}:D{re})')
        ws.cell(ar, 17, f'=MIN(Q{rs}:Q{re})')
        ws.cell(ar, 18, f'=MAX(R{rs}:R{re})')
        ws.cell(ar, 19, f'=IF(OR(Q{ar}=0,R{ar}=0),0,R{ar}-Q{ar})')
        ws.cell(ar, 20, f'=SUM(T{rs}:T{re})')
        ws.cell(ar, 21, f'=IFERROR(T{ar}/(T{S.agg81}+T{S.agg82}),0)')
        ws.cell(ar, 22, f'=U{ar}')
        ws.cell(ar, 23, f'=V{ar}')

    # Checksums
    ws.cell(S.chk0, 3, '✓ ∑ вес ур.3 подзем.')
    ws.cell(S.chk0, 22, f'=SUM(V{S.p0}:V{S.pN})')
    ws.cell(S.chk0 + 1, 3, '✓ ∑ вес ур.3 надзем.')
    ws.cell(S.chk0 + 1, 22, f'=SUM(V{S.n0}:V{S.nN})')
    ws.cell(S.chk0 + 2, 3, '✓ ∑ вес ур.2')
    ws.cell(S.chk0 + 2, 22, f'=V{S.agg81}+V{S.agg82}')
    ws.cell(S.chk3, 3, '✓ ∑ глоб. вес')
    ws.cell(S.chk3, 23, f'=SUM(W{S.p0}:W{S.nN})')

    # Корпуса
    ws.cell(S.korp_hdr, 1, 'РАСПРЕДЕЛЕНИЕ НАДЗЕМНОЙ ЧАСТИ ПО КОРПУСАМ').font = BOLD
    korp_headers = ['Корпус'] + [c[0] for c in NADZEM] + ['Σ план', 'Вес корп.(авто)', 'Вес корп.']
    set_row(ws, S.korp_col, korp_headers, font=BOLD)

    sum_col = 2 + NN
    sum_range = '+'.join(f'{col(sum_col)}{S.k0 + ki}' for ki in range(NK))
    for ki, kname in enumerate(KORPUS):
        r = S.k0 + ki
        ws.cell(r, 1, kname)
        for ci, area in enumerate(KORPUS_AREAS[kname]):
            ws.cell(r, 2 + ci, area)
        last_area = col(2 + NN - 1)
        ws.cell(r, sum_col, f'=SUM(B{r}:{last_area}{r}')
        ws.cell(r, sum_col + 1, f'=IFERROR({col(sum_col)}{r}/({sum_range}),0)')
        ws.cell(r, sum_col + 2, f'={col(sum_col + 1)}{r}')

    ws.cell(S.kchk, 1, '✓ ∑ вес корп.')
    ws.cell(S.kchk, sum_col + 2, f'={"+".join(f"{col(sum_col+2)}{S.k0+ki}" for ki in range(NK))}')

    # Start date
    ws.cell(S.date_r, 1, 'Дата начала проекта')
    ws.cell(S.date_r, 2, START_DATE)
    ws.cell(S.date_r, 2).number_format = DATE_FMT

    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['A'].width = 8


# ================================================================
#  ВВОД
# ================================================================

def build_vvod(wb, sheet_suffix, cats):
    sn = f'ВВОД {sheet_suffix}'
    ws = wb.create_sheet(sn)

    ws.cell(1, 1, f'ВВОД ДАННЫХ — {sn}').font = BOLD
    ws.cell(2, 1, 'Вводите м². Серые столбцы A-D не редактируйте.').font = WARN_FONT
    set_row(ws, 3, ['Дата', 'Тип', 'Код', 'Категория', 'Пол', 'Стены', 'Потолок', 'Мебл.', '✓'], font=BOLD)

    rpw = len(cats) * 2
    date_ref = f'=СПРАВОЧНИК!$B${S.date_r}'

    for w in range(WEEKS):
        ws0 = 4 + w * rpw
        for t_idx, t_name in enumerate(['Черновая', 'Чистовая']):
            for ci, (code, name, *_) in enumerate(cats):
                r = ws0 + t_idx * len(cats) + ci
                if w == 0 and t_idx == 0 and ci == 0:
                    ws.cell(r, 1, date_ref)
                elif t_idx == 0 and ci == 0:
                    prev_first = 4 + (w - 1) * rpw
                    ws.cell(r, 1, f'=A{prev_first}+7')
                else:
                    ws.cell(r, 1, f'=A{ws0}')
                ws.cell(r, 1).number_format = DATE_FMT
                ws.cell(r, 2, t_name)
                ws.cell(r, 3, code)
                ws.cell(r, 4, name)
                ws.cell(r, 9, f'=COUNTA(E{r}:G{r})&"/3"')

    ws.column_dimensions['D'].width = 40


# ================================================================
#  ПЛАН / ФАКТ
# ================================================================

def build_plan_fact(wb, kind, cc, part_suffix, cats, has_mebl_total=False):
    sn = f'{kind} {cc} {part_suffix}'
    ws = wb.create_sheet(sn)

    nc = len(cats)
    gap_c = 2 + nc * 3
    tot_start = gap_c + 1
    n_tot = 4 if has_mebl_total else 3
    tot_labels = ['Пол', 'Стены', 'Потолок'] + (['Мебл.'] if has_mebl_total else [])

    ws.cell(1, 1, sn).font = BOLD
    ws.cell(1, tot_start, f'Итого {part_suffix.lower()}')
    ws.cell(2, 1, 'Дата')
    ws.cell(2, 2, 'Подземная часть' if '8.1' in cats[0][0] else 'Надземная часть').font = BOLD

    for ci, (code, name, *_) in enumerate(cats):
        ws.cell(3, 2 + ci * 3, f'{code} {name}').font = BOLD
    for ci in range(nc):
        for si, sname in enumerate(['Пол', 'Стены', 'Потолок']):
            ws.cell(4, 2 + ci * 3 + si, sname)
    for ti, tl in enumerate(tot_labels):
        ws.cell(3, tot_start + ti, tl).font = BOLD

    date_ref = f'=СПРАВОЧНИК!$B${S.date_r}'
    for w in range(WEEKS):
        r = 5 + w
        ws.cell(r, 1, date_ref if w == 0 else f'=A{r - 1}+7')
        ws.cell(r, 1).number_format = DATE_FMT
        for si in range(min(3, n_tot)):
            src = '+'.join(f'{col(2 + ci * 3 + si)}{r}' for ci in range(nc))
            ws.cell(r, tot_start + si, f'={src}')

    ws.column_dimensions['A'].width = 12


# ================================================================
#  ОТЧЕТ
# ================================================================

def build_otchet(wb):
    ws = wb.create_sheet('ОТЧЕТ')
    print('  ОТЧЕТ')

    ws.cell(1, 1, 'ОТЧЁТ ПО ОТДЕЛОЧНЫМ РАБОТАМ').font = BOLD
    ws.cell(2, 1, '⚠ Автоматический. Не вносите данные напрямую.').font = WARN_FONT

    hdr = ['Код', 'Название', 'Тип', 'Вес ур.4', 'Вес ур.5', 'План м²', 'Факт м²', 'Факт %', 'Взвеш. %']
    last_data_row = 4 + WEEKS

    r = 4

    def write_section(r, title, cats, is_ndz, fact_pfx, plan_fn):
        ws.cell(r, 1, title).font = BOLD
        r += 1
        set_row(ws, r, hdr, font=BOLD)
        r += 1

        for ci, (code, name, plan, wc, wch, cw, chw) in enumerate(cats):
            sr = sprav_cat_row(ci, is_ndz)
            m = r  # main row

            ws.cell(r, 1, code)
            ws.cell(r, 2, f'=СПРАВОЧНИК!C{sr}')
            ws.cell(r, 3, 'Главная')
            ws.cell(r, 6, plan_fn(ci))
            ws.cell(r, 7, f'=G{r+1}*D{r+1}+G{r+5}*D{r+5}')
            ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})')
            ws.cell(r, 8).number_format = PCT_FMT
            ws.cell(r, 9, f'=I{r+1}+I{r+5}')
            ws.cell(r, 9).number_format = PCT_FMT
            r += 1

            # Черновая
            ws.cell(r, 2, 'Черновая')
            ws.cell(r, 4, f'=СПРАВОЧНИК!E{sr}')
            ws.cell(r, 6, f'=F{m}')
            ws.cell(r, 7, f'=G{r+1}*E{r+1}+G{r+2}*E{r+2}+G{r+3}*E{r+3}')
            ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})')
            ws.cell(r, 8).number_format = PCT_FMT
            ws.cell(r, 9, f'=H{r}*D{r}')
            ws.cell(r, 9).number_format = PCT_FMT
            r += 1

            fsc = f'ФАКТ ЧЕРН {fact_pfx}'
            for si, sn in enumerate(['Пол', 'Стены', 'Потолок']):
                fc = fact_col_letter(ci, si)
                ws.cell(r, 2, f'   {sn}')
                ws.cell(r, 3, 'Вид')
                ws.cell(r, 5, f'=СПРАВОЧНИК!{chern_weight_col(si)}{sr}')
                ws.cell(r, 6, f'=F{m}')
                ws.cell(r, 7, f"=SUM('{fsc}'!{fc}5:{fc}{4+WEEKS})")
                ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})')
                ws.cell(r, 8).number_format = PCT_FMT
                ws.cell(r, 9, f'=H{r}*E{r}')
                ws.cell(r, 9).number_format = PCT_FMT
                r += 1

            # Чистовая
            chist_r = r
            ws.cell(r, 2, 'Чистовая')
            ws.cell(r, 4, f'=СПРАВОЧНИК!F{sr}')
            ws.cell(r, 6, f'=F{m}')
            ws.cell(r, 7, f'=G{r+1}*E{r+1}+G{r+2}*E{r+2}+G{r+3}*E{r+3}')
            ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})')
            ws.cell(r, 8).number_format = PCT_FMT
            ws.cell(r, 9, f'=H{r}*D{r}')
            ws.cell(r, 9).number_format = PCT_FMT
            r += 1

            fsi = f'ФАКТ ЧИСТ {fact_pfx}'
            for si, sn in enumerate(['Пол', 'Стены', 'Потолок']):
                fc = fact_col_letter(ci, si)
                ws.cell(r, 2, f'   {sn}')
                ws.cell(r, 3, 'Вид')
                ws.cell(r, 5, f'=СПРАВОЧНИК!{chist_weight_col(si)}{sr}')
                ws.cell(r, 6, f'=F{m}')
                ws.cell(r, 7, f"=SUM('{fsi}'!{fc}5:{fc}{4+WEEKS})")
                ws.cell(r, 8, f'=IF(F{r}=0,0,G{r}/F{r})')
                ws.cell(r, 8).number_format = PCT_FMT
                ws.cell(r, 9, f'=H{r}*E{r}')
                ws.cell(r, 9).number_format = PCT_FMT
                r += 1

            ws.cell(r, 2, '✓ Сумма весов ур.4')
            ws.cell(r, 4, f'=D{m+1}+D{chist_r}')
            r += 2

        return r

    r = write_section(r, 'ПОДЗЕМНАЯ ЧАСТЬ', PODZEM, False, 'ПОДЗЕМ',
                       lambda ci: f'=СПРАВОЧНИК!D{sprav_cat_row(ci, False)}')

    for ki, kname in enumerate(KORPUS):
        kr = S.k0 + ki
        r = write_section(r, f'КОРПУС {kname}', NADZEM, True, kname,
                           lambda ci, _kr=kr: f'=СПРАВОЧНИК!{col(2 + ci)}{_kr}')

    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['A'].width = 8


# ================================================================
#  SIGNAL
# ================================================================

def build_signal(wb):
    ws = wb.create_sheet('SIGNAL')
    print('  SIGNAL')

    blocks = []

    def add_block(label, btype, **kw):
        idx = len(blocks)
        b = {'idx': idx, 'base': 2 + idx * 6, 'label': label, 'type': btype}
        b.update(kw)
        blocks.append(b)
        return b

    # Подземная часть
    bpo = add_block('8.1 Подземная часть: Общий', 'combined')
    bpc = add_block('8.1 Подземная часть: Черновая', 'weighted',
                     sheet='ПЛАН ЧЕРН ПОДЗЕМ', fsheet='ФАКТ ЧЕРН ПОДЗЕМ',
                     cats=PODZEM, is_nadzem=False, wtype='chern')
    bpi = add_block('8.1 Подземная часть: Чистовая', 'weighted',
                     sheet='ПЛАН ЧИСТ ПОДЗЕМ', fsheet='ФАКТ ЧИСТ ПОДЗЕМ',
                     cats=PODZEM, is_nadzem=False, wtype='chist')
    bpo['chern'] = bpc
    bpo['chist'] = bpi
    bpo['vsego'] = f'=СПРАВОЧНИК!D{S.agg81}'

    # Per-корпус
    ko_list, kc_list, ki_list = [], [], []
    for ki_idx, kname in enumerate(KORPUS):
        bko = add_block(f'8.2 {kname}: Общий', 'combined')
        bkc = add_block(f'8.2 {kname}: Черновая', 'weighted',
                         sheet=f'ПЛАН ЧЕРН {kname}', fsheet=f'ФАКТ ЧЕРН {kname}',
                         cats=NADZEM, is_nadzem=True, wtype='chern')
        bki = add_block(f'8.2 {kname}: Чистовая', 'weighted',
                         sheet=f'ПЛАН ЧИСТ {kname}', fsheet=f'ФАКТ ЧИСТ {kname}',
                         cats=NADZEM, is_nadzem=True, wtype='chist')
        bko['chern'] = bkc
        bko['chist'] = bki
        sum_c = 2 + NN
        bko['vsego'] = f'=СПРАВОЧНИК!{col(sum_c)}{S.k0 + ki_idx}'
        ko_list.append(bko)
        kc_list.append(bkc)
        ki_list.append(bki)

    # Надземная агрегация
    bno = add_block('8.2 Надземная часть: Общий', 'sum_blocks', sources=ko_list)
    bnc = add_block('8.2 Надземная часть: Черновая', 'sum_blocks', sources=kc_list)
    bni = add_block('8.2 Надземная часть: Чистовая', 'sum_blocks', sources=ki_list)
    for b in [bno, bnc, bni]:
        b['vsego'] = f'=СПРАВОЧНИК!D{S.agg82}'

    # Общая
    bto = add_block('8. Отделка общий: Общий', 'sum_two', a=bpo, b=bno)
    btc = add_block('8. Отделка общий: Черновая', 'sum_two', a=bpc, b=bnc)
    bti = add_block('8. Отделка общий: Чистовая', 'sum_two', a=bpi, b=bni)
    for b in [bto, btc, bti]:
        b['vsego'] = f'=СПРАВОЧНИК!D{S.agg81}+СПРАВОЧНИК!D{S.agg82}'

    # Report date
    ws.cell(1, 1, 'Дата отчета')
    ws.cell(2, 1, REPORT_DATE)
    ws.cell(2, 1).number_format = DATE_FMT

    # Write blocks
    for b in blocks:
        bc = b['base']
        dc = bc + 1
        pc = bc + 2
        fc = bc + 3

        # Metadata (strings!)
        ws.cell(1, bc, 'Заголовок'); ws.cell(1, dc, 'План-факт по объемам')
        ws.cell(2, bc, 'Статус');    ws.cell(2, dc, 'true')
        ws.cell(3, bc, 'Url изображения')
        ws.cell(4, bc, 'Тип');      ws.cell(4, dc, 'planFact2')
        ws.cell(6, bc, 'Дата');      ws.cell(6, dc, '=$A$2')
        ws.cell(7, bc, 'Гистограмма'); ws.cell(7, dc, 'false')
        ws.cell(8, bc, 'По месяцам');  ws.cell(8, dc, 'false')
        ws.cell(9, bc, 'Всего');     ws.cell(9, dc, b.get('vsego', ''))
        ws.cell(10, bc, 'Тип');      ws.cell(10, dc, b['label'])
        ws.cell(11, bc, 'Ед. изм.'); ws.cell(11, dc, 'м2')
        ws.cell(12, bc, 'Учитывать дату карточки'); ws.cell(12, dc, 'true')
        ws.cell(13, bc, 'Данные')

        # Data rows
        for w in range(WEEKS):
            r = 14 + w
            pr = 5 + w

            if w == 0:
                ws.cell(r, dc, f'=СПРАВОЧНИК!$B${S.date_r}')
            else:
                ws.cell(r, dc, f'={col(dc)}{r - 1}+7')
            ws.cell(r, dc).number_format = DATE_FMT

            bt = b['type']

            if bt == 'weighted':
                nc = len(b['cats'])
                is_ndz = b['is_nadzem']
                wcol_fn = chist_weight_col if b['wtype'] == 'chist' else chern_weight_col

                plan_t = []
                fact_t = []
                for ci in range(nc):
                    sr = sprav_cat_row(ci, is_ndz)
                    for si in range(3):
                        pcl = fact_col_letter(ci, si)
                        wc = wcol_fn(si)
                        plan_t.append(f"'{b['sheet']}'!{pcl}{pr}*СПРАВОЧНИК!${wc}${sr}")
                        fact_t.append(f"'{b['fsheet']}'!{pcl}{pr}*СПРАВОЧНИК!${wc}${sr}")
                ws.cell(r, pc, '=' + '+'.join(plan_t))
                ws.cell(r, fc, '=' + '+'.join(fact_t))

            elif bt == 'combined':
                chp = b['chern']['base'] + 2
                chip = b['chist']['base'] + 2
                chf = b['chern']['base'] + 3
                chif = b['chist']['base'] + 3
                ws.cell(r, pc, f'={col(chp)}{r}+{col(chip)}{r}')
                ws.cell(r, fc, f'={col(chf)}{r}+{col(chif)}{r}')

            elif bt == 'sum_blocks':
                srcs = b['sources']
                ws.cell(r, pc, '=' + '+'.join(f'{col(s["base"]+2)}{r}' for s in srcs))
                ws.cell(r, fc, '=' + '+'.join(f'{col(s["base"]+3)}{r}' for s in srcs))

            elif bt == 'sum_two':
                ws.cell(r, pc, f'={col(b["a"]["base"]+2)}{r}+{col(b["b"]["base"]+2)}{r}')
                ws.cell(r, fc, f'={col(b["a"]["base"]+3)}{r}+{col(b["b"]["base"]+3)}{r}')

    # Named ranges for SIGNAL blocks
    sig_ranges = []
    for b in blocks:
        bc = b['base']
        c1 = col(bc)
        c2 = col(bc + 3)

        label = b['label']
        if '8.1' in label:
            pfx = ''
        else:
            pfx = ''
        # Generate name from label
        name = _signal_range_name(b)
        sig_ranges.append((name, f"'SIGNAL'!${c1}:${c2}"))

    return ws, sig_ranges


def _signal_range_name(b):
    label = b['label']
    for kn in KORPUS:
        if kn in label:
            if 'Общий' in label: return f'{kn}_Общий_8_2'
            if 'Черновая' in label: return f'{kn}_Черновая_8_2'
            if 'Чистовая' in label: return f'{kn}_Чистовая_8_2'
    if '8.1' in label:
        if 'Общий' in label: return 'Общий_8_1'
        if 'Черновая' in label: return 'Черновая_8_1'
        if 'Чистовая' in label: return 'Чистовая_8_1'
    if '8.2' in label and 'Надземная' in label:
        if 'Общий' in label: return 'Общий_8_2'
        if 'Черновая' in label: return 'Черновая_8_2'
        if 'Чистовая' in label: return 'Чистовая_8_2'
    if 'Отделка общий' in label:
        if 'Общий' in label: return 'Общая'
        if 'Черновая' in label: return 'Черновая'
        if 'Чистовая' in label: return 'Чистовая'
    return label.replace(' ', '_').replace(':', '').replace('.', '_')


# ================================================================
#  NAMED RANGES
# ================================================================

def add_named_ranges(wb, signal_ranges):
    print('  Named ranges (SIGNAL only)')
    count = 0

    for name, ref in signal_ranges:
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names.add(dn)
        count += 1

    print(f'    {count} total')


# ================================================================
#  MAIN
# ================================================================

def generate():
    print(f'Generating: {NK} корпусов ({", ".join(KORPUS)})')
    expected = 4 * NK + 7
    print(f'Expected sheets: {expected}\n')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_spravochnik(wb)

    print('  ВВОД sheets')
    build_vvod(wb, 'ПОДЗЕМ', PODZEM)
    for k in KORPUS:
        build_vvod(wb, k, NADZEM)

    print('  ПЛАН/ФАКТ sheets')
    all_parts = [('ПОДЗЕМ', PODZEM, False)] + [(k, NADZEM, True) for k in KORPUS]
    for suffix, cats, is_ndz in all_parts:
        for kind in ['ПЛАН', 'ФАКТ']:
            build_plan_fact(wb, kind, 'ЧЕРН', suffix, cats)
            build_plan_fact(wb, kind, 'ЧИСТ', suffix, cats, has_mebl_total=is_ndz)

    build_otchet(wb)
    ws_sig, sig_ranges = build_signal(wb)
    wb.create_sheet('Лист2')
    add_named_ranges(wb, sig_ranges)

    print(f'\nTotal sheets: {len(wb.sheetnames)}')
    print(f'Saving: {OUTPUT}')
    wb.save(OUTPUT)
    print('DONE!')


if __name__ == '__main__':
    generate()
